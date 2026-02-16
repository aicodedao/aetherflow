from __future__ import annotations
from pathlib import Path
from typing import Any, Dict
from aetherflow.core.api import Step, register_step

@register_step("demo_prepare_excel_template")
class DemoPrepareExcelTemplate(Step):
    required_inputs = {"template_path"}
    def run(self) -> Dict[str, Any]:
        self.validate()
        try:
            import openpyxl
        except Exception as e:
            raise RuntimeError("openpyxl is required (install aetherflow-core[excel] or [reports])") from e

        job_dir = self.ctx.artifacts_dir(self.job_id)
        tp = Path(job_dir / self.inputs["template_path"]).resolve()
        tp.parent.mkdir(parents=True, exist_ok=True)

        if tp.exists():
            return {"template_path": str(tp), "created": False}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        ws["A1"] = "Sales"
        ws["A10"] = "Cost"

        wb.defined_names.append(openpyxl.workbook.defined_name.DefinedName("SALES_ANCHOR", attr_text="Report!$A$2"))
        wb.defined_names.append(openpyxl.workbook.defined_name.DefinedName("COST_ANCHOR", attr_text="Report!$A$11"))

        wb.save(tp)
        return {"template_path": str(tp), "created": True}
