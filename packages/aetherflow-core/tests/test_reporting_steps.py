import json
import sqlite3
from pathlib import Path

import pytest

from aetherflow.core.context import RunContext
from aetherflow.core.runtime.settings import Settings
from aetherflow.core.connectors.base import ConnectorInit
from aetherflow.core.builtins.connectors import SQLiteDB
from aetherflow.core.builtins.steps import (
    DbFetchSmall,
    DbExtractStream,
    ExcelFillSmall,
    ExcelFillFromFile,
    ReportTooLargeError,
)
from aetherflow.core.steps._io import fast_count_rows


def _make_ctx(tmp: Path, settings: Settings, env={}):
    work_root = Path(settings.work_root)
    work_root.mkdir(parents=True, exist_ok=True)
    return RunContext(
        settings=settings,
        flow_id="test_flow",
        run_id="run1",
        work_root=work_root,
        layout={"artifacts": "artifacts", "scratch": "scratch", "manifests": "manifests"},
        state=None,
        resources={},
        env=env,
        connectors={},
    )


def _make_sqlite_db(path: Path, *, rows: int):
    conn = sqlite3.connect(path)
    try:
        cur = conn.cursor()
        cur.execute("create table t(id integer primary key, v text)")
        cur.executemany("insert into t(v) values (?)", [(f"v{i}",) for i in range(rows)])
        conn.commit()
    finally:
        conn.close()


def test_db_fetch_small_limit_triggers(temp_dir, settings):
    db_path = temp_dir / "t.sqlite"
    _make_sqlite_db(db_path, rows=6)

    ctx = _make_ctx(temp_dir, settings)
    init = ConnectorInit(name="db", kind="db", driver="sqlite3", config={"path": str(db_path)}, options={}, ctx=None)
    ctx.connectors["db_main"] = SQLiteDB(init)

    step = DbFetchSmall(
        "fetch", {"resource": "db_main", "sql": "select id, v from t order by id", "max_rows": 5}, ctx, "job"
    )
    with pytest.raises(ValueError):
        step.run()


def test_db_extract_stream_writes_file_and_rowcount(temp_dir, settings):
    db_path = temp_dir / "t.sqlite"
    _make_sqlite_db(db_path, rows=3)

    ctx = _make_ctx(temp_dir, settings)
    init = ConnectorInit(name="db", kind="db", driver="sqlite3", config={"path": str(db_path)}, options={}, ctx=None)
    ctx.connectors["db_main"] = SQLiteDB(init)

    step = DbExtractStream(
        "extract",
        {"resource": "db_main", "sql": "select id, v from t order by id", "output": "out.tsv", "format": "tsv"},
        ctx,
        "job",
    )
    out = step.run()
    assert out["row_count"] == 3
    p = Path(out["artifact_path"])
    assert p.exists()
    txt = p.read_text(encoding="utf-8").strip().splitlines()
    assert txt[0] == "id	v"
    assert len(txt) == 4


def test_db_extract_stream_file_options_delimiter_and_linefeed(temp_dir, settings):
    db_path = temp_dir / "t.sqlite"
    _make_sqlite_db(db_path, rows=2)

    ctx = _make_ctx(temp_dir, settings)
    init = ConnectorInit(name="db", kind="db", driver="sqlite3", config={"path": str(db_path)}, options={}, ctx=None)
    ctx.connectors["db_main"] = SQLiteDB(init)

    step = DbExtractStream(
        "extract",
        {
            "resource": "db_main",
            "sql": "select id, v from t order by id",
            "output": "out.csv",
            "format": "csv",
            "delimiter": "|",
            "linefeed": "\r\n",
            "encoding": "utf-8",
        },
        ctx,
        "job",
    )
    out = step.run()
    p = Path(out["artifact_path"])
    raw = p.read_bytes()
    assert b"id|v\r\n" in raw


def test_excel_fill_small_and_overlap_validation(temp_dir, settings):
    pytest.importorskip("openpyxl")
    import openpyxl
    from openpyxl.workbook.defined_name import DefinedName

    ctx = _make_ctx(temp_dir, settings)
    tpl = Path(ctx.settings.work_root) / "tpl.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["B2"].value = "ANCHOR"
    wb.defined_names.add(DefinedName("tbl_anchor", attr_text="Report!$B$2"))
    wb.save(tpl)

    rows = [["r1c1", "r1c2"], ["r2c1", "r2c2"]]
    step = ExcelFillSmall(
        "fill",
        {
            "template_path": str(tpl),
            "output": "out.xlsx",
            "targets": [
                {"name": "t1", "anchor": "tbl_anchor", "rows_json": json.dumps(rows), "max_rows": 10, "max_cols": 10},
            ],
        },
        ctx,
        "job",
    )
    out = step.run()
    out_xlsx = Path(out["output"])
    assert out_xlsx.exists()

    wb2 = openpyxl.load_workbook(out_xlsx)
    ws2 = wb2["Report"]
    assert ws2["B2"].value == "r1c1"
    assert ws2["C3"].value == "r2c2"

    # Overlap validation
    step2 = ExcelFillSmall(
        "fill2",
        {
            "template_path": str(tpl),
            "output": "out2.xlsx",
            "targets": [
                {"name": "a", "sheet": "Report", "cell": "B2", "rows_json": json.dumps([["x"]])},
                {"name": "b", "sheet": "Report", "cell": "B2", "rows_json": json.dumps([["y"]])},
            ],
        },
        ctx,
        "job",
    )
    with pytest.raises(ValueError):
        step2.run()


def test_excel_fill_from_file_data_sheet_and_threshold_guard(temp_dir, settings):
    pytest.importorskip("openpyxl")
    import openpyxl

    tpl = temp_dir / "tpl.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Report"
    wb.save(tpl)

    tsv = temp_dir / "data.tsv"
    tsv.write_text("amount\tday\n10\t2026-02-08\n20\t2026-02-09\n", encoding="utf-8")

    ctx = _make_ctx(temp_dir, settings, env = {"AETHERFLOW_STRICT_SANDBOX": "false"})

    # data_sheet mode creates DATA_ sheet
    step = ExcelFillFromFile(
        "x",
        {
            "template_path": str(tpl),
            "output": "report.xlsx",
            "targets": [
                {
                    "name": "sales",
                    "source_path": str(tsv),
                    "source_format": "tsv",
                    "mode": "data_sheet",
                    "data_sheet_prefix": "DATA_",
                    "cell": "A1",
                }
            ],
        },
        ctx,
        "job",
    )
    out = step.run()
    out_xlsx = Path(out["output"])
    wb2 = openpyxl.load_workbook(out_xlsx)
    assert "DATA_sales" in wb2.sheetnames
    ws = wb2["DATA_sales"]
    assert ws["A1"].value == "amount"
    assert ws["B3"].value == "2026-02-09"

    # report_region mode enforces threshold (count-before-write)
    step2 = ExcelFillFromFile(
        "y",
        {
            "template_path": str(tpl),
            "output": "report2.xlsx",
            "rows_threshold": 1,
            "targets": [
                {
                    "name": "sales",
                    "sheet": "Report",
                    "cell": "A1",
                    "source_path": str(tsv),
                    "source_format": "tsv",
                    "mode": "report_region",
                    "fail_on_threshold": True,
                }
            ],
        },
        ctx,
        "job",
    )
    with pytest.raises(ReportTooLargeError):
        step2.run()


def test_fast_count_rows_csv_parse_handles_multiline_quoted_fields(temp_dir):
    # One data row, but contains an embedded newline inside quotes.
    p = temp_dir / "m.csv"
    p.write_text('a,b\n1,"hello\nworld"\n', encoding="utf-8")

    # Fast mode miscounts (two lines of data). csv_parse returns correct 1.
    fast = fast_count_rows(p, "csv", include_header=True, count_mode="fast", linefeed="\n", encoding="utf-8")
    parsed = fast_count_rows(p, "csv", include_header=True, count_mode="csv_parse", encoding="utf-8")
    assert parsed == 1
    assert fast != parsed


def test_db_extract_stream_emit_dtypes(temp_dir, settings):
    db_path = temp_dir / "t.sqlite"
    _make_sqlite_db(db_path, rows=2)

    ctx = _make_ctx(temp_dir, settings)
    init = ConnectorInit(name="db", kind="db", driver="sqlite3", config={"path": str(db_path)}, options={}, ctx=None)
    ctx.connectors["db_main"] = SQLiteDB(init)

    step = DbExtractStream(
        "extract",
        {
            "resource": "db_main",
            "sql": "select id, v from t order by id",
            "output": "out.tsv",
            "format": "tsv",
            "emit_dtypes": True,
        },
        ctx,
        "job",
    )
    out = step.run()
    assert "dtypes" in out
    assert out["dtypes"].get("id") == "int"
    assert out["dtypes"].get("v") == "string"


def test_excel_fill_from_file_typed_cast_csv_tsv(temp_dir, settings):
    pytest.importorskip("openpyxl")
    import openpyxl

    tpl = temp_dir / "tpl.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Report"
    wb.save(tpl)

    tsv = temp_dir / "typed.tsv"
    tsv.write_text("amount\tday\n10\t2026-02-08\n20\t2026-02-09\n", encoding="utf-8")

    ctx = _make_ctx(temp_dir, settings, env = {"AETHERFLOW_STRICT_SANDBOX": "false"})

    step = ExcelFillFromFile(
        "fill",
        {
            "template_path": str(tpl),
            "output": "typed.xlsx",
            "targets": [
                {
                    "name": "t",
                    "source_path": str(tsv),
                    "source_format": "tsv",
                    "mode": "data_sheet",
                    "cell": "A1",
                    "type_cast": "schema",
                    "dtypes": {"amount": "int", "day": "date"},
                }
            ],
        },
        ctx,
        "job",
    )
    out = step.run()
    wb2 = openpyxl.load_workbook(Path(out["output"]))
    ws = wb2["DATA_t"]
    assert ws["A2"].value == 10
    assert isinstance(ws["A2"].value, int)
    assert str(ws["B2"].value).startswith("2026-02-08")


def test_fast_count_rows_csv_parse_handles_multiline_csv(temp_dir):
    csv_p = temp_dir / "m.csv"
    csv_p.write_text('id,comment\n1,"hello\nworld"\n', encoding="utf-8")

    # fast mode assumes one record per line (likely miscounts here)
    n_fast = fast_count_rows(csv_p, "csv", include_header=True, count_mode="fast")
    n_parse = fast_count_rows(csv_p, "csv", include_header=True, count_mode="csv_parse")

    assert n_parse == 1
    assert n_fast != n_parse
