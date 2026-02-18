from __future__ import annotations

import re
import datetime as _dt
import fnmatch
import hashlib
import json
import logging
import os
import posixpath
import shutil
import signal
import subprocess
import threading
import time
from dataclasses import dataclass, replace, field
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, Tuple, List, Optional, Iterable

from aetherflow.core.concurrency import run_thread_pool
from aetherflow.core.exception import ConnectorError, ParquetSupportMissing, ReportTooLargeError
from aetherflow.core.observability import log_event
from aetherflow.core.registry.steps import get_step, register_step
from aetherflow.core.spec import RemoteFileMeta
from aetherflow.core.steps._io import fast_count_rows
from aetherflow.core.steps.base import Step, StepResult, STEP_SKIPPED, STEP_SUCCESS

log = logging.getLogger("aetherflow.core.builtin.steps")


def _reject_symlink_chain(root: Path, cand: Path) -> None:
    """
    Reject any symlink in path chain root -> cand (inclusive).
    This inspects path segments WITHOUT resolving them so that symlinks
    inside the chain are reliably detected.
    """
    root = Path(root)
    cand = Path(cand)
    try:
        # both must be absolute or made absolute on caller
        root = root.resolve()  # ok to resolve root once
    except Exception:
        raise ValueError(f"Invalid sandbox root: {root}")
    # Build the candidate path *without* resolving symlinks (preserve original path components)
    # Get the relative parts of candidate against root, but using pure path math:
    try:
        rel = cand.relative_to(root)
    except Exception:
        # If cand is not under root, reject upstream caller should already handle this,
        # but we keep a clear error.
        raise ValueError(f"Path escapes artifacts_dir: {cand}")
    cur = root
    # iterate each path segment from root to cand (inclusive)
    for part in rel.parts:
        cur = cur / part
        # if the segment exists on filesystem, check if the segment itself is a symlink
        # Use lstat/is_symlink on the path itself (do NOT resolve)
        try:
            if cur.exists() and cur.is_symlink():
                raise ValueError(f"Path contains symlink (sandbox blocked): {cur}")
        except OSError:
            # If we can't stat the path, be conservative and block
            raise ValueError(f"Unable to validate path segment for symlink: {cur}")


def _allowed_roots(ctx, job_id: str) -> list[Path]:
    roots: list[Path] = []
    settings = getattr(ctx, "settings", None)
    # work_root (always allowed)
    if settings and getattr(settings, "work_root", None):
        roots.append(Path(settings.work_root).resolve())
    # job artifacts root always allowed
    try:
        # artifacts root (parent of job artifacts)
        artifacts_root = Path(ctx.artifacts_dir(job_id)).resolve()
        roots.append(artifacts_root)
    except Exception:
        pass
    # dedupe
    return list({r for r in roots if r.exists()})


def _resolve_path(ctx, job_id: str, p: str) -> Path:
    """
    Resolve user-provided paths.

    - relative paths → resolved against job artifacts dir
    - absolute paths → allowed only if under ALLOWED_ROOTS (when sandbox=True)
    """
    raw = str(p)
    path = Path(raw)
    artifacts_root = Path(ctx.artifacts_dir(job_id)).resolve()
    if path.is_absolute():
        cand = path.resolve()
    else:
        cand = (artifacts_root / path).resolve()
    if not _is_sandbox(ctx):
        return cand
    # Enterprise mode: enforce strict containment
    if _is_enterprise(ctx):
        try:
            cand.relative_to(artifacts_root)
        except Exception:
            raise ValueError(
                f"Path escapes job artifacts dir in enterprise mode: {raw}"
            )
        _reject_symlink_chain(artifacts_root, cand)
        return cand
    # Non-enterprise mode: allow under ALLOWED_ROOTS
    allowed = _allowed_roots(ctx, job_id)
    for root in allowed:
        try:
            cand.relative_to(root)
            _reject_symlink_chain(root, cand)
            return cand
        except Exception:
            continue
    raise ValueError(
        f"Path not under allowed roots: {raw}. "
        f"Allowed roots: {[str(r) for r in allowed]}"
    )


def _is_enterprise(ctx) -> bool:
    env = getattr(ctx, "env", None)
    if isinstance(env, dict):
        return str(env.get("AETHERFLOW_MODE_ENTERPRISE", "false")).lower() in ("1", "true", "yes", "y", "on")
    # fallback: object with attrs
    return bool(getattr(env, "AETHERFLOW_MODE_ENTERPRISE", False) if env is not None else False)


def _is_sandbox(ctx) -> bool:
    env = getattr(ctx, "env", None)
    if isinstance(env, dict):
        return str(env.get("AETHERFLOW_STRICT_SANDBOX", "true")).lower() in ("1", "true", "yes", "y", "on")
    # fallback: object with attrs
    return bool(getattr(env, "AETHERFLOW_STRICT_SANDBOX", True) if env is not None else False)


def _as_list(x: Any) -> list:
    if x is None:
        return []
    if isinstance(x, list):
        return x
    return [x]


@dataclass
class _Capture:
    mode: str
    max_bytes: int
    file_path: Optional[Path] = None
    _buf: bytearray = field(default_factory=bytearray, repr=False)  # type: ignore

    def __post_init__(self):
        if self._buf is None:
            self._buf = bytearray()

    def add(self, data: bytes) -> None:
        if self.mode != "capture":
            return
        if len(self._buf) >= self.max_bytes:
            return
        remain = self.max_bytes - len(self._buf)
        self._buf.extend(data[:remain])

    def text(self) -> str:
        try:
            return self._buf.decode("utf-8", errors="replace")
        except Exception:
            return ""


def _stream_reader(pipe, cap: _Capture, logger, *, settings, event_prefix: str, fields: dict, level: int):
    try:
        for line in iter(pipe.readline, b""):
            cap.add(line)
            # emit line-by-line events (keeps json logs parseable)
            msg = line.decode("utf-8", errors="replace").rstrip("\n")
            if msg and logger is not None and settings is not None:
                log_event(logger, settings=settings, level=level, event=event_prefix, message=msg, **fields)
    finally:
        try:
            pipe.close()
        except Exception:
            pass


def _check_success(*, ctx, job_id: str, spec: dict) -> Tuple[bool, str | None]:
    """Return (ok, reason)."""
    require_files = [str(x) for x in _as_list(spec.get("require_files"))]
    require_glob = [str(x) for x in _as_list(spec.get("require_glob"))]
    forbid_files = [str(x) for x in _as_list(spec.get("forbid_files"))]

    marker = spec.get("marker_file")
    if marker:
        require_files.append(str(marker))

    for p in require_files:
        rp = _resolve_path(ctx, job_id, p)
        if not rp.exists():
            return False, f"missing_required_file:{p}"

    for g in require_glob:
        # glob is evaluated relative to the filesystem root if absolute, otherwise relative to artifacts dir
        base = _resolve_path(ctx, job_id, g)
        matches = list(base.parent.glob(base.name))
        if not matches:
            return False, f"missing_required_glob:{g}"

    for p in forbid_files:
        rp = _resolve_path(ctx, job_id, p)
        if rp.exists():
            return False, f"forbidden_file_present:{p}"

    return True, None


@register_step("db_extract")
class DbExtract(Step):
    required_inputs = {"resource", "sql", "output"}

    def run(self) -> Dict[str, Any]:
        self.validate()
        db = self.ctx.connectors[self.inputs["resource"]]
        sql = self.inputs["sql"]
        params = self.inputs.get("params")
        fmt = (self.inputs.get("format") or "tsv").lower()

        out_path = _resolve_path(self.ctx, self.job_id, self.inputs["output"])
        out_path.parent.mkdir(parents=True, exist_ok=True)
        tmp = out_path.with_suffix(out_path.suffix + ".tmp")

        cols, rows = db.fetchall(sql, params)

        if fmt == "tsv":
            with open(tmp, "w", encoding="utf-8") as f:
                if cols:
                    f.write("\t".join(cols) + "\n")
                for r in rows:
                    f.write("\t".join("" if v is None else str(v) for v in r) + "\n")
        elif fmt == "jsonl":
            with open(tmp, "w", encoding="utf-8") as f:
                for r in rows:
                    obj = {cols[i]: r[i] for i in range(len(cols))}
                    f.write(json.dumps(obj, ensure_ascii=False) + "\n")
        else:
            raise ValueError(f"Unsupported format: {fmt}")

        os.replace(tmp, out_path)
        return {"output": str(out_path), "format": fmt, "rows": len(rows)}


#######################################################################################

def _maybe_json_load(x: Any) -> Any:
    """Best-effort JSON parse for templated inputs.

    Aetherflow renders templates as strings; for steps that accept large
    structured inputs we support passing *JSON strings*.
    """
    if not isinstance(x, str):
        return x
    s = x.strip()
    if not s:
        return x
    if (s.startswith("[") and s.endswith("]")) or (s.startswith("{") and s.endswith("}")):
        try:
            return json.loads(s)
        except Exception:
            return x
    return x


def _parse_jsonish(x: Any) -> Any:
    """Parse JSON if x is a JSON string, otherwise return x unchanged."""
    return _maybe_json_load(x)


def _file_opt(d: dict, key: str, default: Any = None) -> Any:
    """Read a file-format option from either top-level or nested `file:` dict.

    This keeps inputs backwards compatible while allowing a structured style:

      file:
        encoding: utf-8
        delimiter: "\t"
    """

    if key in d and d[key] is not None:
        return d[key]
    fd = d.get("file") or {}
    if isinstance(fd, dict) and key in fd and fd[key] is not None:
        return fd[key]
    return default


def _db_fetchmany_unified(db, sql: str, params: dict | None, *, fetch_size: int, sample_size: int = 200):
    """
    Returns: (cols: list[str], it: Iterator[tuple], pytypes: list[type])
    Contract:
      - If connector implements fetchmany(), use it.
      - Else ConnectorError.
    """
    if hasattr(db, "fetchmany") and callable(getattr(db, "fetchmany")):
        cols, it, pytypes = db.fetchmany(sql, params, fetch_size=fetch_size, sample_size=sample_size)
        return cols or [], it, pytypes or [object for _ in (cols or [])]
    raise ConnectorError("DB-Connector does not have built-in fetchmany!")


def _pytype_to_dtype(t: type) -> str:
    if t is bool:
        return "bool"
    if t is int:
        return "int"
    if t is float:
        return "float"
    if t is Decimal:
        # if no support decimal in parquet, type map -> downgrade float
        log.warning("parquet dtype: Decimal downgraded to float (precision loss possible)")
        return "float"
    if t is _dt.date:
        return "date"
    if t is _dt.datetime:
        return "datetime"
    return "string"


def _json_safe(v: Any):
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (_dt.date, _dt.datetime)):
        return v.isoformat()
    return v


@register_step("db_fetch_small")
class DbFetchSmall(Step):
    """Fetch a small result set into memory (guarded).

    Because Aetherflow templates render to strings, downstream steps should
    consume the JSON-string outputs (`rows_json`, `columns_json`) when chaining.

    Inputs:
      - resource: resource name
      - sql: query
      - params: optional dict
      - max_rows: int (default 50_000)
      - fetch_size: int (default 5000)
    """

    required_inputs = {"resource", "sql"}

    def run(self) -> Dict[str, Any]:
        self.validate()
        db = self.ctx.connectors[self.inputs["resource"]]
        sql = self.inputs["sql"]
        params = self.inputs.get("params")
        max_rows = int(self.inputs.get("max_rows", 50_000))
        fetch_size = int(self.inputs.get("fetch_size", 5000))

        cols, it, pytypes = _db_fetchmany_unified(db, sql, params, fetch_size=fetch_size)
        pytypes = (pytypes or [])
        if cols and len(pytypes) != len(cols):
            pytypes = (pytypes + [object] * len(cols))[: len(cols)]
        dtypes = {str(c): _pytype_to_dtype(t) for c, t in zip(cols, pytypes)}

        rows: list[list[Any]] = []
        count = 0
        for r in it:
            rows.append([None if v is None else v for v in r])
            count += 1
            if count > max_rows:
                raise ValueError(f"db_fetch_small exceeded max_rows={max_rows}. Use db_extract_stream for large results.")

        rows_safe = [[_json_safe(v) for v in r] for r in rows]
        return {
            "columns": cols,
            "rows": rows,
            "row_count": count,
            "columns_json": json.dumps(cols, ensure_ascii=False),
            "rows_json": json.dumps(rows_safe, ensure_ascii=False),
            #"pytypes": [t.__name__ for t in pytypes],
            #"pytypes_json": json.dumps([t.__name__ for t in pytypes]),
            "dtypes": dtypes,
            "dtypes_json": json.dumps(dtypes, ensure_ascii=False)
        }


@register_step("db_extract_stream")
class DbExtractStream(Step):

    required_inputs = {"resource", "sql", "output"}

    def run(self) -> Dict[str, Any]:
        self.validate()

        db = self.ctx.connectors[self.inputs["resource"]]
        sql = self.inputs["sql"]
        params = self.inputs.get("params")
        fmt = (self.inputs.get("format") or "tsv").lower()
        fetch_size = int(self.inputs.get("fetch_size", 5000))
        include_header = bool(self.inputs.get("include_header", True))

        emit_dtypes = bool(self.inputs.get("emit_dtypes", False))
        dtypes = _parse_jsonish(self.inputs.get("dtypes")) or _parse_jsonish(self.inputs.get("dtypes_json"))
        if dtypes is not None and not isinstance(dtypes, dict):
            raise ValueError("dtypes must be a dict (mapping column -> type)")

        out_path = _resolve_path(self.ctx, self.job_id, self.inputs["output"])
        out_path.parent.mkdir(parents=True, exist_ok=True)
        tmp = out_path.with_suffix(out_path.suffix + ".tmp")

        cols, it, pytypes = _db_fetchmany_unified(
            db, sql, params, fetch_size=fetch_size
        )

        pytypes = (pytypes or [])
        if cols and len(pytypes) != len(cols):
            pytypes = (pytypes + [object] * len(cols))[: len(cols)]

        if emit_dtypes and not dtypes:
            dtypes = {str(c): _pytype_to_dtype(t) for c, t in zip(cols, pytypes)}

        row_count = 0

        try:
            if fmt in ("tsv", "csv"):
                import csv
                encoding = str(_file_opt(self.inputs, "encoding", "utf-8"))
                delimiter = str(_file_opt(self.inputs, "delimiter", "\t" if fmt == "tsv" else ","))
                quotechar = str(_file_opt(self.inputs, "quotechar", '"'))
                escapechar = _file_opt(self.inputs, "escapechar", None)
                doublequote = bool(_file_opt(self.inputs, "doublequote", True))
                quoting = _file_opt(self.inputs, "quoting", "minimal")
                linefeed = str(_file_opt(self.inputs, "linefeed", "\n"))
                q = {
                    "minimal": csv.QUOTE_MINIMAL,
                    "all": csv.QUOTE_ALL,
                    "none": csv.QUOTE_NONE,
                    "nonnumeric": csv.QUOTE_NONNUMERIC,
                }.get(str(quoting).strip().lower(), csv.QUOTE_MINIMAL)
                with open(tmp, "w", encoding=encoding, newline="") as f:
                    w = csv.writer(
                        f,
                        delimiter=delimiter,
                        quotechar=quotechar,
                        escapechar=escapechar,
                        doublequote=doublequote,
                        quoting=q,
                        lineterminator=linefeed,
                    )
                    if include_header and cols:
                        w.writerow(cols)
                    for r in it:
                        w.writerow(["" if v is None else str(v) for v in r])
                        row_count += 1

            elif fmt == "parquet":
                try:
                    import pyarrow as pa
                    import pyarrow.parquet as pq
                except Exception as e:
                    raise ValueError(
                        "parquet format requires optional dependency: pyarrow "
                        "(install aetherflow-core[parquet])"
                    ) from e

                def _pa_type(t: str):
                    t = (t or "string").lower()
                    if t == "int":
                        return pa.int64()
                    if t == "float":
                        return pa.float64()
                    if t == "bool":
                        return pa.bool_()
                    if t == "date":
                        return pa.date32()
                    if t == "datetime":
                        return pa.timestamp("ms")
                    return pa.string()

                dtype_map = {str(k): str(v) for k, v in (dtypes or {}).items()}
                schema = pa.schema(
                    [(c, _pa_type(dtype_map.get(str(c), "string"))) for c in cols]
                )
                writer = None
                batch_rows = []
                try:
                    writer = pq.ParquetWriter(tmp, schema=schema)
                    for r in it:
                        batch_rows.append(r)
                        row_count += 1

                        if len(batch_rows) >= fetch_size:
                            arrays = []
                            for i, c in enumerate(cols):
                                pa_t = schema.field(i).type
                                col_vals = [
                                    row[i] if i < len(row) else None
                                    for row in batch_rows
                                ]
                                if not pa.types.is_string(pa_t):
                                    col_vals = [None if v == "" else v for v in col_vals]
                                arrays.append(pa.array(col_vals, type=pa_t))
                            writer.write_table(pa.Table.from_arrays(arrays, names=cols))
                            batch_rows = []
                    if batch_rows:
                        arrays = []
                        for i, c in enumerate(cols):
                            pa_t = schema.field(i).type
                            col_vals = [
                                row[i] if i < len(row) else None
                                for row in batch_rows
                            ]
                            if not pa.types.is_string(pa_t):
                                col_vals = [None if v == "" else v for v in col_vals]
                            arrays.append(pa.array(col_vals, type=pa_t))
                        writer.write_table(pa.Table.from_arrays(arrays, names=cols))
                finally:
                    if writer is not None:
                        writer.close()
            else:
                raise ValueError(f"db_extract_stream unsupported format: {fmt}")

            os.replace(tmp, out_path)

        finally:
            # Cleanup temp if something failed before replace
            try:
                if tmp.exists():
                    tmp.unlink()
            except Exception:
                pass

        # sha256
        h = hashlib.sha256()
        with open(out_path, "rb") as f:
            for b in iter(lambda: f.read(1024 * 1024), b""):
                h.update(b)

        result = {
            "artifact_path": str(out_path),
            "format": fmt,
            "columns": cols,
            "row_count": row_count,
            "sha256": h.hexdigest(),
        }

        if emit_dtypes and dtypes:
            result["dtypes"] = dtypes
            result["dtypes_json"] = json.dumps(dtypes, ensure_ascii=False)

        return result


##################################################
### Excel Helpers
##################################################

from copy import copy as _copy_obj

def _excel_require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except Exception as e:
        raise ValueError("Excel steps require optional dependency: openpyxl (install aetherflow-core[excel])") from e

def _rect_intersects(a, b) -> bool:
    (r1, c1, r2, c2) = a
    (s1, d1, s2, d2) = b
    return not (r2 < s1 or s2 < r1 or c2 < d1 or d2 < c1)

def _excel_resolve_anchor(wb, *, sheet_name: str | None, anchor: str | None, cell: str | None, **kwargs):
    """
    Resolve an Excel target to (worksheet, start_row, start_col).

    Supported:
      - anchor = named range
      - anchor = "Sheet!A1" or "A1"
      - anchor = "QIP_ANCHOR" / "QC_ANCHOR" (anchor text search, robust normalize)
      - cell   = "Sheet!A1" or "A1"
    """
    from openpyxl.utils.cell import coordinate_to_tuple, column_index_from_string
    import re

    def _norm(s: str) -> str:
        # robust normalize: strip, replace NBSP, collapse whitespace
        s = "" if s is None else str(s)
        s = s.replace("\u00A0", " ")  # NBSP -> space
        s = s.strip()
        s = re.sub(r"\s+", " ", s)
        return s

    def _norm_key(s: str) -> str:
        # for matching (case-insensitive)
        return _norm(s).upper()

    def _is_a1(s: str) -> bool:
        s = _norm(s).replace("$", "").upper()
        return bool(re.match(r"^[A-Z]{1,3}[0-9]{1,7}$", s))

    def _parse_sheet_a1(s: str) -> tuple[str | None, str]:
        s = _norm(s)
        if "!" in s:
            sh, c = s.split("!", 1)
            return _norm(sh) or None, _norm(c)
        return None, s

    def _resolve_named_range(name: str):
        dn = wb.defined_names.get(name)
        if dn is None:
            return None
        for title, coord in dn.destinations:
            try:
                ws = wb[title]
            except Exception:
                continue
            coord0 = str(coord).split(":", 1)[0].replace("$", "").strip().upper()
            m = re.match(r"^([A-Z]+)([0-9]+)$", coord0)
            if not m:
                continue
            col_letters, row_s = m.group(1), m.group(2)
            return ws, int(row_s), int(column_index_from_string(col_letters))
        return None

    def _resolve_cell_ref(sh: str | None, a1: str):
        a1 = _norm(a1).replace("$", "").upper()
        if not _is_a1(a1):
            raise ValueError(f"Invalid Excel cell reference: {a1}")
        if sh is None:
            if not sheet_name:
                raise ValueError("Excel target requires sheet when using cell like 'A1' (set sheet or use 'Sheet!A1')")
            ws = wb[sheet_name]
        else:
            ws = wb[sh]
        r, c = coordinate_to_tuple(a1)
        return ws, int(r), int(c)

    def _search_text(text: str, *, only_sheet: str | None, max_cells: int = 10000, enterprise_mode: bool = False):
        needle = _norm_key(text)
        if not needle:
            return None
        if enterprise_mode:
            # enterprise: MUST pin to a sheet to avoid O(N) scan
            if not only_sheet:
                raise ValueError("Excel anchor text search requires 'sheet' in enterprise mode. Use named range or set sheet.")
        scanned = 0
        sheets = [wb[only_sheet]] if only_sheet and only_sheet in wb.sheetnames else list(wb.worksheets)
        for ws in sheets:
            for row in ws.iter_rows(values_only=False):
                for cell_obj in row:
                    scanned += 1
                    if scanned > max_cells:
                        raise ValueError(f"Excel anchor search exceeded max_cells={max_cells}. Use named range or specify sheet.")
                    v = cell_obj.value
                    if v is None:
                        continue
                    try:
                        sv = str(v)  # <-- covers rich text objects too
                    except Exception:
                        continue
                    if _norm_key(sv) == needle:
                        return ws, int(cell_obj.row), int(cell_obj.column)
        return None

    # --------------------
    # Resolution order
    # --------------------
    if anchor:
        raw = kwargs.get("log_level", "INFO")
        if isinstance(raw, int):
            log_level = raw
        else:
            log_level = getattr(logging, str(raw).upper(), logging.INFO)

        a = _norm(anchor)
        # 1) Named range
        hit = _resolve_named_range(a)
        if hit is not None:
            return hit
        # 2) If anchor looks like Sheet!A1 or A1
        sh, maybe_a1 = _parse_sheet_a1(a)
        if _is_a1(maybe_a1):
            return _resolve_cell_ref(sh, maybe_a1)
        # 3) Anchor text search (prefer provided sheet_name)
        hit = _search_text(a, only_sheet=sheet_name, enterprise_mode=kwargs.get("enterprise_mode", False))
        if hit is not None:
            return hit
        # DEBUG: show what workbook actually contains (helps detect wrong template / hidden chars)
        found = []
        if log_level <= logging.DEBUG:
            try:
                for ws2 in wb.worksheets:
                    for row in ws2.iter_rows(values_only=False):
                        for co in row:
                            v = co.value
                            if v is None:
                                continue
                            try:
                                sv = str(v)
                            except Exception:
                                continue
                            if "ANCHOR" in sv.upper():
                                found.append(f"{ws2.title}!{co.coordinate}={sv!r}")
                            if len(found) >= 20:
                                raise StopIteration
            except StopIteration:
                pass
            except Exception:
                pass
        raise ValueError(
            f"Excel anchor not found: {a!r}. "
            f"Workbook sheets={wb.sheetnames}. "
            f"ANCHOR-like cells(sample)={found}"
        )

    if not cell:
        raise ValueError("Excel target requires either anchor (named range / anchor text) or cell")
    c = _norm(cell)
    sh, a1 = _parse_sheet_a1(c)
    return _resolve_cell_ref(sh, a1)

def _parse_iso_date(s: str):
    try:
        return _dt.date.fromisoformat(s)
    except Exception:
        return None

def _parse_iso_datetime(s: str):
    try:
        if s.endswith("Z"):
            s = s[:-1]
        return _dt.datetime.fromisoformat(s)
    except Exception:
        return None

def _cast_value(v: Any, dtype: str):
    if v is None:
        return None
    if isinstance(v, (int, float, bool, _dt.date, _dt.datetime)):
        return v
    s = str(v)
    if s == "":
        return None
    t = (dtype or "string").lower()
    if t == "string":
        return s
    if t == "int":
        try:
            return int(s)
        except Exception:
            return s
    if t == "float":
        try:
            return float(s)
        except Exception:
            return s
    if t == "bool":
        sl = s.strip().lower()
        if sl in ("1", "true", "t", "yes", "y"):
            return True
        if sl in ("0", "false", "f", "no", "n"):
            return False
        return s
    if t == "date":
        d = _parse_iso_date(s)
        return d if d is not None else s
    if t == "datetime":
        d = _parse_iso_datetime(s)
        return d if d is not None else s
    return s

# cast helpers
def _cast_row(row, *, columns, type_cast, schema_map):
    if type_cast == "none":
        return row
    if type_cast == "basic":
        out = []
        for v in row:
            s = "" if v is None else str(v)
            if s == "":
                out.append(None)
                continue
            sl = s.strip().lower()
            if sl in ("true", "false", "1", "0"):
                out.append(_cast_value(s, "bool"))
                continue
            if s.isdigit() and (s == "0" or not s.startswith("0")):
                out.append(_cast_value(s, "int"))
                continue
            try:
                if "." in s or "e" in sl:
                    out.append(float(s))
                    continue
            except Exception:
                pass
            d = _parse_iso_datetime(s)
            if d is not None:
                out.append(d)
                continue
            d2 = _parse_iso_date(s)
            if d2 is not None:
                out.append(d2)
                continue
            out.append(s)
        return out

    if not schema_map or not columns:
        return row
    out = []
    for idx, v in enumerate(row):
        col = columns[idx] if idx < len(columns) else None
        dt = schema_map.get(str(col), "string") if col is not None else "string"
        out.append(_cast_value(v, dt))
    return out

def _excel_copy_cell_style(src_cell, dst_cell) -> None:
    """
    Copy visual formatting from src_cell to dst_cell (openpyxl-safe).
    Best-effort: styling must not break data writing.
    """
    try:
        dst_cell.font = _copy_obj(src_cell.font)
        dst_cell.fill = _copy_obj(src_cell.fill)
        dst_cell.border = _copy_obj(src_cell.border)
        dst_cell.alignment = _copy_obj(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = _copy_obj(src_cell.protection)
    except Exception:
        pass

def _excel_build_style_row_cache(ws, style_row: int, start_col: int, width: int):
    """
    Cache style cells for a row so we don't look up per-write.
    """
    return [ws.cell(row=style_row, column=start_col + j) for j in range(width)]

def _excel_set_cell_value(ws, r: int, c: int, v: Any = None):
    cell = ws.cell(row=r, column=c)
    cell.value = v
    return cell

def _excel_clear_row_segment(ws, row: int, col0: int, width: int) -> None:
    """
    Clear values across a row segment (keeps formatting).
    This is the "clean override" behavior for template headers.
    """
    if width <= 0:
        return
    for j in range(width):
        try:
            _excel_set_cell_value(ws=ws, r=row, c=col0+j)
        except Exception:
            pass

def _excel_clear_row_values(ws, row: int, start_col: int, width: int) -> None:
    """Clear values on a row (keeps formatting)."""
    _excel_clear_row_segment(ws, row, start_col, width)


def _iter_target_sheets(wb, sheet_name: str | None):
    """
    Return list of worksheet objects to operate on.
    If sheet_name is provided it must exist in workbook.
    """
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        return [wb[sheet_name]]
    return list(wb.worksheets)


def _norm_excel_text(s: str) -> str:
    s = "" if s is None else str(s)
    s = s.replace("\u00A0", " ")
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    return s


@register_step("excel_validate_template")
class ExcelValidateTemplate(Step):
    required_inputs = {"template_path", "required_names"}

    def run(self) -> Dict[str, Any]:
        self.validate()
        _excel_require_openpyxl()
        import openpyxl

        template_path = _resolve_path(self.ctx, self.job_id, self.inputs["template_path"])
        log.info("excel_validate_template template_path=%s", str(template_path))

        wb = openpyxl.load_workbook(template_path, data_only=True)
        # ---- 1) Collect named ranges ----
        defined_names = set(wb.defined_names.keys())
        # ---- 2) Collect anchor text values in all sheets ----
        anchor_values = set()
        wss = _iter_target_sheets(wb=wb, sheet_name=self.inputs.get("sheet"))
        for ws in wss:
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if isinstance(v, str):
                        anchor_values.add(_norm_excel_text(v))
        # ---- 3) Normalize required list ----
        req = self.inputs["required_names"]
        if isinstance(req, str):
            try:
                req = json.loads(req)
            except Exception:
                req = [req]
        if not isinstance(req, list):
            raise ValueError(f"required_names must be a list of strings {req}")
        missing = []
        found_named = []
        found_anchor = []
        for name in req:
            if name in defined_names:
                found_named.append(name)
            elif _norm_excel_text(name) in anchor_values:
                found_anchor.append(name)
            else:
                missing.append(name)
        if missing:
            raise ValueError(f"Template missing named ranges or anchor cells: {missing} {req}")
        return {
            "template_ok": True,
            "found_named_ranges": sorted(found_named),
            "found_anchor_cells": sorted(found_anchor),
        }


@register_step("excel_fill_small")
class ExcelFillSmall(Step):
    """
    Fill an Excel template with small in-memory tables (JSON rows).

    This step is intended for *small* datasets that fit comfortably in memory
    (typically produced by `db_fetch_small`). It writes rows into a pre-formatted
    Excel template, optionally preserving and/or overriding template headers, and
    optionally copying styles from a template "style row".

    For large datasets from files (CSV/TSV/Parquet) that should be streamed into
    Excel without loading everything into memory, use `excel_fill_from_file`.

    =============================================================================
    Targeting / Anchor Resolution
    =============================================================================
    Each target MUST set exactly ONE of:
      - anchor: a named range OR an anchor text cell (exact match after normalize)
      - cell:   a direct cell reference ("A1" when sheet is provided) or "Sheet!A1"

    Anchor resolution order (when using `anchor`) is implemented by `_excel_resolve_anchor`:
      1) Named range (defined name)
      2) If anchor looks like "Sheet!A1" or "A1", treat it as a cell reference
      3) Otherwise search for a cell whose normalized text equals the anchor token

    =============================================================================
    Template Layout Model
    =============================================================================
    After resolving the anchor/cell, we get (ws, r_anchor, c0).

    Region start row `r0` is derived from `anchor_is_marker`:

      - anchor_is_marker=True  => r0 = r_anchor + 1
          Use this when the anchor row is a marker-only row:
              [ANCHOR]              <- marker
              [HEADER TEMPLATE...]  <- r0
              [DATA...]             <- r0+1

      - anchor_is_marker=False => r0 = r_anchor
          Use this when the anchor lives on the header row:
              [ANCHOR][HEADER...]   <- r0
              [DATA...]             <- r0+1

    =============================================================================
    HARD-LOCKED Insert Semantics (prevents header-style bugs)
    =============================================================================
    insert controls how we avoid overwriting content below:

      - insert="replace": write in-place (overwrites existing cell values)
      - insert="below":   insert rows before writing to preserve content below

    The critical question is whether the template already contains a header row
    at r0 that must be preserved (position + formatting). This is controlled by:

      template_has_header (bool)

    HARD-LOCK rules:
      - If template_has_header=True:
          * Data starts at data_start_row = r0 + 1
          * For insert="below": we insert ONLY at data_start_row (never at r0)
            so the template header row at r0 is never pushed down.
      - If template_has_header=False:
          * Data starts at r0 unless we also write a header (header_mode override/append)
          * For insert="below": we insert the full block at the block start row

    This guarantees the tricky case:
      insert=below + header_style=template + header_mode=template
    still inserts in the correct place (below the template header), preserving
    header formatting and preventing duplicate headers.

    =============================================================================
    Header Writing
    =============================================================================
    Header behavior depends on:
      - header_mode: "template" | "override" | "append"
      - columns_json: optional list[str]

    Rules:
      - If header_mode="template": we do NOT write header values (template header stays).
      - If header_mode in ("override","append") and columns_json is provided:
          * We write header values at r0.
          * If header_clear=True, we clear old header VALUES across header_clear_width
            before writing (formatting remains intact).
      - header_style controls header formatting only when header_mode writes a header:
          * header_style="template": keep existing header formatting at r0.
          * header_style="style_row": copy formatting from style_row onto header cells
            (requires style_mode=copy_row and style_apply includes header/both).

    =============================================================================
    Type Casting
    =============================================================================
    Targets can apply type casting before writing to Excel:

      - type_cast: "none" | "basic" | "schema"
      - dtypes / dtypes_json: dict[str,str] mapping column name -> dtype

    Supported dtypes: "string", "int", "float", "bool", "date", "datetime"

    Rules:
      - type_cast="none":   do not modify values.
      - type_cast="basic": best-effort casting from strings (bool/int/float/date/datetime).
      - type_cast="schema": cast by column name using schema_map. Requires columns_json
        (without columns_json, schema casting is a no-op and the row is returned as-is).

    Implementation uses module helpers:
      - `_cast_value`
      - `_cast_row(row, columns=..., type_cast=..., schema_map=...)`

    =============================================================================
    Styling (copy row)
    =============================================================================
    Style copying is optional and controlled by:

      - style_mode: "none" | "copy_row"
      - style_row_offset: int (default 2)
      - style_apply: "data" | "header" | "both"
      - clear_style_row: bool

    style_row index:
      - if anchor_is_marker=True:  style_row = r_anchor + style_row_offset
      - else:                      style_row = r0 + style_row_offset

    IMPORTANT header rule:
      - If header_style="template", header will NOT be styled from style_row even if
        style_apply includes header (template formatting is preserved).
      - If header_style="style_row", header styling follows style_apply.

    =============================================================================
    Anchor clearing (optional)
    =============================================================================
    Some templates include visible marker text (e.g. "QIP_ANCHOR") that should be removed
    after locating the region. Use:

      - anchor_clear: bool
      - anchor_clear_mode: "cell" | "row"
      - anchor_clear_width: int (only for "row")

    Clearing uses `_excel_set_cell_value` / `_excel_clear_row_segment` to keep formatting.

    =============================================================================
    Inputs (step-level)
    =============================================================================
    template_path (str, required)
    output (str, required)
    targets (list[object], required)

    =============================================================================
    Target schema (targets[*])
    =============================================================================
    name (str, optional)
    sheet (str, optional)
    anchor (str, optional) XOR cell (str, optional)
    rows_json (str, required)
    columns_json (str, optional)
    row_count (int, optional)

    header_mode ("template"|"override"|"append", default "template")
    header_style ("template"|"style_row", default "template")
    header_clear (bool, default True when header_mode writes header)
    header_clear_width (int, optional)

    insert ("below"|"replace", default "replace")
    anchor_is_marker (bool, default True if anchor else False)
    template_has_header (bool, default True if anchor else False)

    anchor_clear (bool, default False)
    anchor_clear_mode ("cell"|"row", default "cell")
    anchor_clear_width (int, optional)

    type_cast ("none"|"basic"|"schema", default "schema" if dtypes provided else "none")
    dtypes / dtypes_json (dict, optional)

    style_mode ("none"|"copy_row", default "none")
    style_row_offset (int, default 2)
    style_apply ("data"|"header"|"both", default "data")
    clear_style_row (bool, default False)

    max_rows (int, optional)
        Guard for DATA rows only (does not include header row).
    max_total_rows (int, optional)
        Guard for total written rows (header + data), if desired.
    max_cols (int, optional)

    =============================================================================
    Outputs
    =============================================================================
    {"output": str, "written": list[...]}
    """

    required_inputs = {"template_path", "output", "targets"}

    def run(self) -> Dict[str, Any]:
        self.validate()
        _excel_require_openpyxl()
        import openpyxl

        template_path = _resolve_path(self.ctx, self.job_id, self.inputs["template_path"])
        log.info("excel_fill_small template_path=%s", str(template_path))

        out_path = _resolve_path(self.ctx, self.job_id, self.inputs["output"])
        out_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.load_workbook(template_path)
        targets = self.inputs.get("targets") or []

        # Overlap guard is PER-SHEET: (sheet_name, r1, c1, r2, c2)
        rects: list[tuple[str, int, int, int, int]] = []
        written: list[dict[str, Any]] = []

        for t in targets:
            name = t.get("name") or "target"

            row_count = t.get("row_count")
            # ----------------------------
            # Guards
            # ----------------------------
            max_rows = t.get("max_rows")
            if max_rows is not None and row_count is not None and int(row_count) > int(max_rows):
                raise ValueError(f"excel_fill_small target={name}: row_count={row_count} exceed max_rows={max_rows}")
            max_total_rows = t.get("max_total_rows")
            if max_total_rows is not None and row_count is not None and int(row_count) > int(max_total_rows):
                raise ValueError(f"excel_fill_small target={name}: row_count={row_count} exceed max_total_rows={max_total_rows}")

            sheet = t.get("sheet")
            anchor = t.get("anchor")
            cell = t.get("cell")
            if bool(anchor) == bool(cell):
                raise ValueError(f"excel_fill_small target={name}: must set exactly one of anchor OR cell")

            rows = _maybe_json_load(t.get("rows_json"))
            cols = _maybe_json_load(t.get("columns_json")) if t.get("columns_json") is not None else None
            if isinstance(rows, str):
                raise ValueError(f"excel_fill_small target={name}: rows_json must be valid JSON")
            if cols is not None and isinstance(cols, str):
                raise ValueError(f"excel_fill_small target={name}: columns_json must be valid JSON")

            ws, r_anchor, c0 = _excel_resolve_anchor(
                wb,
                sheet_name=sheet,
                anchor=anchor,
                cell=cell,
                log_level=self.ctx.settings.log_level,
                enterprise_mode=_is_enterprise(self.ctx)
            )

            insert = str(t.get("insert") or "replace").lower()
            if insert not in ("below", "replace"):
                raise ValueError(f"excel_fill_small target={name}: invalid insert={insert}")

            header_mode = str(t.get("header_mode") or "template").lower()
            if header_mode not in ("template", "override", "append"):
                raise ValueError(f"excel_fill_small target={name}: invalid header_mode={header_mode}")

            header_style = str(t.get("header_style") or "template").lower()
            if header_style not in ("template", "style_row"):
                raise ValueError(f"excel_fill_small target={name}: invalid header_style={header_style}")

            # Determine anchor kind (named range or not) for sane defaults
            is_named_anchor = False
            if anchor:
                try:
                    is_named_anchor = anchor in wb.defined_names
                except Exception:
                    is_named_anchor = False
            # anchor_is_marker default:
            # - named range => False (write starts exactly at that cell)
            # - text anchor => True (common pattern: marker row then header row below)
            if "anchor_is_marker" in t:
                anchor_is_marker = bool(t.get("anchor_is_marker"))
            else:
                anchor_is_marker = False if is_named_anchor else (True if anchor else False)

            r0 = r_anchor + (1 if anchor_is_marker else 0)

            # HARD-LOCK defaults
            if "template_has_header" in t:
                template_has_header = bool(t.get("template_has_header"))
            else:
                # only assume template header exists when anchor row is a marker
                template_has_header = True if (anchor and anchor_is_marker) else False

            # Optional: clear anchor marker text (after resolve, before insert/write)
            if bool(t.get("anchor_clear", False)) and anchor:
                mode_clear = str(t.get("anchor_clear_mode") or "cell").lower()
                if mode_clear not in ("cell", "row"):
                    raise ValueError(f"excel_fill_small target={name}: invalid anchor_clear_mode={mode_clear}")
                if mode_clear == "cell":
                    _excel_set_cell_value(ws=ws, r=r_anchor, c=c0)
                else:
                    raw_w = t.get("anchor_clear_width", None)
                    try:
                        w_clear = int(raw_w) if raw_w is not None else 1
                    except Exception:
                        w_clear = 1
                    if w_clear > 0:
                        _excel_clear_row_segment(ws, r_anchor, c0, w_clear)

            # ----------------------------
            # Header + casting setup
            # ----------------------------
            data_rows_raw: list[list[Any]] = [list(x) for x in (rows or [])]
            header_row = list(cols) if cols else None
            columns = header_row

            # dtypes/schema
            dtypes = _parse_jsonish(t.get("dtypes")) or _parse_jsonish(t.get("dtypes_json"))
            if dtypes is not None and not isinstance(dtypes, dict):
                raise ValueError(f"excel_fill_small target={name}: dtypes must be a dict")
            schema_map = {str(k): str(v) for k, v in (dtypes or {}).items()}

            type_cast = (t.get("type_cast") or ("schema" if schema_map else "none")).lower()
            if type_cast not in ("none", "schema", "basic"):
                raise ValueError(f"excel_fill_small target={name}: invalid type_cast={type_cast}")

            # Apply casting BEFORE any insert/write
            data_rows: list[list[Any]] = [
                _cast_row(r, columns=columns, type_cast=type_cast, schema_map=schema_map)
                for r in data_rows_raw
            ]

            will_write_header = (header_mode in ("override", "append")) and (header_row is not None)

            # Determine where DATA starts (even when we do not write a header)
            reserve_template_header_row = bool(template_has_header)
            data_start_row = r0 + (1 if (will_write_header or reserve_template_header_row) else 0)

            # Rows written (conceptual)
            write_rows: list[list[Any]] = []
            if will_write_header:
                write_rows.append(header_row)
            write_rows.extend(data_rows)

            # ----------------------------
            # Guards
            # ----------------------------
            if max_rows is not None and len(data_rows) > int(max_rows):
                raise ValueError(f"excel_fill_small target={name}: data rows exceed max_rows={max_rows}")
            if max_total_rows is not None and len(write_rows) > int(max_total_rows):
                raise ValueError(f"excel_fill_small target={name}: total rows exceed max_total_rows={max_total_rows}")
            width = max((len(r) for r in write_rows), default=0)
            max_cols = t.get("max_cols")
            if max_cols is not None and width > int(max_cols):
                raise ValueError(f"excel_fill_small target={name}: cols exceed max_cols={max_cols}")

            # ----------------------------
            # Insert rows before writing
            # ----------------------------
            if insert == "below":
                if reserve_template_header_row:
                    # HARD-LOCK: never insert at r0 when template header exists
                    n_data = len(data_rows)
                    if n_data > 0:
                        ws.insert_rows(data_start_row, amount=n_data)
                else:
                    if len(write_rows) > 0:
                        ws.insert_rows(r0, amount=len(write_rows))

            # ----------------------------
            # Styling config
            # ----------------------------
            style_mode = str(t.get("style_mode") or "none").lower()
            if style_mode not in ("none", "copy_row"):
                raise ValueError(f"excel_fill_small target={name}: invalid style_mode={style_mode}")

            style_apply = str(t.get("style_apply") or "data").lower()
            if style_apply not in ("data", "header", "both"):
                raise ValueError(f"excel_fill_small target={name}: invalid style_apply={style_apply}")

            style_row_offset = int(t.get("style_row_offset", 2))
            clear_style_row = bool(t.get("clear_style_row", False))

            style_cache = None
            style_row_idx = None
            if style_mode == "copy_row" and width > 0:
                style_row_idx = (r_anchor + style_row_offset) if anchor_is_marker else (r0 + style_row_offset)
                style_cache = _excel_build_style_row_cache(ws, style_row_idx, c0, width)

            # If override/append header, optionally clear old header VALUES but keep formatting
            if will_write_header and header_mode in ("override", "append"):
                header_clear = bool(t.get("header_clear", True))
                header_clear_width = t.get("header_clear_width", None)

                w_new = len(header_row)
                w_guess = width
                if header_clear_width is None:
                    w_clear = max(w_new, w_guess)
                else:
                    try:
                        w_clear = int(header_clear_width)
                    except Exception:
                        w_clear = max(w_new, w_guess)

                if header_clear and w_clear > 0:
                    _excel_clear_row_segment(ws, r0, c0, w_clear)

            # ----------------------------
            # Overlap guard (per-sheet)
            # ----------------------------
            total_rows = (1 if (will_write_header or reserve_template_header_row) else 0) + len(data_rows)
            rect = (ws.title, r0, c0, r0 + max(0, total_rows - 1), c0 + max(0, width - 1))
            for prev in rects:
                prev_sheet, pr1, pc1, pr2, pc2 = prev
                if prev_sheet != ws.title:
                    continue
                if _rect_intersects((pr1, pc1, pr2, pc2), rect[1:]):
                    raise ValueError(f"excel_fill_small targets overlap: name={name} sheet={ws.title} rect={rect} prev={prev}")

            rects.append(rect)

            # ----------------------------
            # Write header (if any) at r0
            # IMPORTANT: do NOT rely on ws.cell(..., value=...)
            # ----------------------------
            if will_write_header:
                for j, v in enumerate(header_row):
                    dst = _excel_set_cell_value(ws=ws, r=r0, c=c0+j, v=v)
                    if style_cache is not None:
                        # apply to header only if requested AND header_style=style_row
                        if header_style == "style_row" and style_apply in ("header", "both"):
                            if j < len(style_cache):
                                _excel_copy_cell_style(style_cache[j], dst)

            # ----------------------------
            # Write data rows at data_start_row
            # ----------------------------
            for i, row in enumerate(data_rows):
                for j, v in enumerate(row):
                    dst = _excel_set_cell_value(ws=ws, r=data_start_row+i, c=c0+j, v=v)
                    if style_cache is not None and style_apply in ("data", "both"):
                        if j < len(style_cache):
                            _excel_copy_cell_style(style_cache[j], dst)

            if style_cache is not None and clear_style_row and style_row_idx is not None:
                _excel_clear_row_values(ws, style_row_idx, c0, width)

            written.append(
                {
                    "name": name,
                    "sheet": ws.title,
                    "rows": int(total_rows),
                    "cols": int(width),
                    "insert": insert,
                    "header_mode": header_mode,
                    "header_style": header_style,
                    "type_cast": type_cast,
                    "style_mode": style_mode,
                    "start_cell": (r0, c0),
                    "template_has_header": bool(template_has_header),
                }
            )

        tmp = out_path.with_suffix(out_path.suffix + ".tmp")
        wb.save(tmp)
        os.replace(tmp, out_path)
        return {"output": str(out_path), "written": written}


@register_step("excel_fill_from_file")
class ExcelFillFromFile(Step):
    """
    Stream CSV/TSV/Parquet data into an Excel template.

    Modes:
      - data_sheet: write a raw export table into a dedicated sheet.
      - report_region: write into a pre-formatted report area using anchor/cell.

    --------------------------------------------------------------------
    Mode support summary (what applies where)
    --------------------------------------------------------------------
    Applies to BOTH modes (data_sheet + report_region):
      - template_path, output, targets
      - anchor XOR cell resolution (if provided)
      - header_mode/header_style + header_clear/header_clear_width
      - style_mode/style_apply/style_row_offset/clear_style_row
      - type_cast + dtypes_json + columns_json
      - anchor_is_marker
      - anchor_clear / anchor_clear_mode / anchor_clear_width
      - template_has_header (hard-lock insert + write placement)

    report_region ONLY:
      - insert (below/replace; default below)
      - rows_threshold / fail_on_threshold + count_mode (guards large regions)

    data_sheet ONLY:
      - data_sheet_prefix (sheet naming)
      - freeze_panes / autofilter / header_bold niceties

    --------------------------------------------------------------------
    HARD-LOCKED template header model
    --------------------------------------------------------------------
    The most common report template layout is:

        [ANCHOR]              <- marker row (anchor_is_marker=True)
        [HEADER TEMPLATE...]  <- r0
        [DATA...]             <- r0+1

    Or:

        [ANCHOR][HEADER...]   <- r0 (anchor_is_marker=False)
        [DATA...]             <- r0+1

    The step must know whether r0 contains a template header row that must be
    preserved (style + row position). This is controlled by:

      template_has_header (bool)

    Defaults (HARD-LOCK):
      - mode=report_region: template_has_header=True
      - mode=data_sheet:    template_has_header=False

    If template_has_header=True:
      - Data ALWAYS starts at data_start_row = r0 + 1
      - For insert="below", rows are inserted at data_start_row (never at r0),
        so the template header row is never pushed down.
      - This specifically fixes the bug scenario:
            insert=below + header_style=template + header_mode=template
        where we are not writing a header, but we MUST insert data under the
        existing header template without destroying its formatting.

    If template_has_header=False:
      - Data starts at r0 when header is not written.
      - insert="below" inserts at the block start row as usual.

    --------------------------------------------------------------------
    Key options (targets[*]) — detailed
    --------------------------------------------------------------------
    mode ("data_sheet"|"report_region", optional, default "data_sheet")

    source_path (str, required), source_format (csv|tsv|parquet)

    sheet (str, optional)
      - data_sheet: optional (auto-created)
      - report_region: required unless anchor is named range or cell includes sheet

    anchor (str) XOR cell (str)
      - anchor: named range or anchor text cell
      - cell: "A1" (requires sheet) or "Sheet!A1"

    insert ("below"|"replace", optional)
      - report_region default: below
      - data_sheet default: replace

    anchor_is_marker (bool, optional)
      - report_region + anchor default: True
      - otherwise default: False

    template_has_header (bool, optional)
      - report_region default: True
      - data_sheet default: False

    header_mode ("template"|"override"|"append")
      - report_region default: template
      - data_sheet default: append

    header_style ("template"|"style_row", default "template")
      - template: keep template formatting on header row (clear values then write)
      - style_row: overwrite header formatting from style row (requires style_mode=copy_row)

    header_clear/header_clear_width
      - only used when header_mode is override/append AND a header is written

    anchor_clear (bool), anchor_clear_mode ("cell"|"row"), anchor_clear_width
      - optional cleanup of visible anchor marker text

    style_mode ("none"|"copy_row"), style_row_offset, style_apply ("data"|"header"|"both")
      - style_apply affects header only when header_style="style_row"

    =============================================================================
    Type Casting
    =============================================================================
    Targets can apply type casting before writing to Excel:

      - type_cast: "none" | "basic" | "schema"
      - dtypes / dtypes_json: dict[str,str] mapping column name -> dtype

    Supported dtypes: "string", "int", "float", "bool", "date", "datetime"

    Rules:
      - type_cast="none":   do not modify values.
      - type_cast="basic": best-effort casting from strings (bool/int/float/date/datetime).
      - type_cast="schema": cast by column name using schema_map. Requires columns_json
        (without columns_json, schema casting is a no-op and the row is returned as-is).

    Implementation uses module helpers:
      - `_cast_value`
      - `_cast_row(row, columns=..., type_cast=..., schema_map=...)`

    --------------------------------------------------------------------
    Outputs
    --------------------------------------------------------------------
    {"output": str, "written": list[...]}
    """

    required_inputs = {"template_path", "output", "targets"}

    def run(self) -> Dict[str, Any]:
        self.validate()
        _excel_require_openpyxl()
        import openpyxl
        import csv

        template_path = _resolve_path(self.ctx, self.job_id, self.inputs["template_path"])
        log.info("excel_fill_from_file template_path=%s", str(template_path))

        out_path = _resolve_path(self.ctx, self.job_id, self.inputs["output"])
        out_path.parent.mkdir(parents=True, exist_ok=True)

        default_rows_threshold = int(self.inputs.get("rows_threshold", 50_000))

        wb = openpyxl.load_workbook(template_path)
        rects: list[tuple[str, int, int, int, int]] = []
        # format: (sheet_name, r1, c1, r2, c2)
        written: list[dict[str, Any]] = []

        for t in (self.inputs.get("targets") or []):
            name = t.get("name") or "target"
            source_path = _resolve_path(self.ctx, self.job_id, t.get("source_path"))
            fmt = (t.get("source_format") or source_path.suffix.lstrip(".") or "tsv").lower()
            mode = (t.get("mode") or "data_sheet").lower()
            if mode not in ("data_sheet", "report_region"):
                raise ValueError(f"excel_fill_from_file target={name}: invalid mode={mode}")

            # legacy include_header support
            if "include_header" in t:
                inc = bool(t.get("include_header", True))
                read_header = inc
                header_mode = "append" if inc else "template"
            else:
                read_header = bool(t.get("read_header", True))
                header_mode = str(
                    t.get("header_mode")
                    or ("template" if mode == "report_region" else "append")
                ).lower()

            if header_mode not in ("template", "override", "append"):
                raise ValueError(f"excel_fill_from_file target={name}: invalid header_mode={header_mode}")

            header_style = str(t.get("header_style") or "template").lower()
            if header_style not in ("template", "style_row"):
                raise ValueError(f"excel_fill_from_file target={name}: invalid header_style={header_style}")

            # File read options
            encoding = str(_file_opt(t, "encoding", "utf-8"))
            delimiter = _file_opt(t, "delimiter", None)
            if delimiter is None and fmt in ("tsv", "csv"):
                delimiter = "\t" if fmt == "tsv" else ","
            delimiter = str(delimiter) if delimiter is not None else ","
            quotechar = str(_file_opt(t, "quotechar", '"'))
            escapechar = _file_opt(t, "escapechar", None)
            doublequote = bool(_file_opt(t, "doublequote", True))
            quoting = _file_opt(t, "quoting", "minimal")
            linefeed = str(_file_opt(t, "linefeed", "\n"))
            count_mode = str(t.get("count_mode", "fast")).lower()

            provided_row_count = t.get("row_count")
            if isinstance(provided_row_count, str) and provided_row_count.strip().isdigit():
                provided_row_count = int(provided_row_count.strip())
            elif isinstance(provided_row_count, (int, float)):
                provided_row_count = int(provided_row_count)
            else:
                provided_row_count = None

            rows_threshold = int(t.get("rows_threshold", default_rows_threshold))

            dtypes = _parse_jsonish(t.get("dtypes")) or _parse_jsonish(t.get("dtypes_json"))
            if dtypes is not None and not isinstance(dtypes, dict):
                raise ValueError(f"excel_fill_from_file target={name}: dtypes must be a dict")
            schema_map = {str(k): str(v) for k, v in (dtypes or {}).items()}

            type_cast = (t.get("type_cast") or ("schema" if schema_map else "none")).lower()
            if type_cast not in ("none", "schema", "basic"):
                raise ValueError(f"excel_fill_from_file target={name}: invalid type_cast={type_cast}")

            provided_columns = _parse_jsonish(t.get("columns")) or _parse_jsonish(t.get("columns_json"))
            if provided_columns is not None and not isinstance(provided_columns, list):
                raise ValueError(f"excel_fill_from_file target={name}: columns must be a list")

            # resolve sheet
            sheet_name = t.get("sheet")
            anchor = t.get("anchor")
            cell = t.get("cell")
            if mode == "data_sheet":
                prefix = str(t.get("data_sheet_prefix") or "DATA_")
                if not sheet_name:
                    sheet_name = f"{prefix}{name}"
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)
                # default start cell A1 if user didn't specify anchor/cell
                if not anchor and not cell:
                    cell = "A1"
            if anchor or cell:
                if bool(anchor) == bool(cell):
                    raise ValueError(f"excel_fill_from_file target={name}: must set exactly one of anchor OR cell")

            ws, r_anchor, c0 = _excel_resolve_anchor(
                wb,
                sheet_name=sheet_name,
                anchor=anchor,
                cell=cell,
                log_level=self.ctx.settings.log_level,
                enterprise_mode=_is_enterprise(self.ctx)
            )

            insert = str(t.get("insert") or ("below" if mode == "report_region" else "replace")).lower()
            if insert not in ("below", "replace"):
                raise ValueError(f"excel_fill_from_file target={name}: invalid insert={insert}")

            # Determine anchor kind for sane defaults
            is_named_anchor = False
            if anchor:
                try:
                    is_named_anchor = anchor in wb.defined_names
                except Exception:
                    is_named_anchor = False

            if "anchor_is_marker" in t:
                anchor_is_marker = bool(t.get("anchor_is_marker"))
            else:
                # named range => usually points to start cell (NOT marker)
                if anchor and is_named_anchor:
                    anchor_is_marker = False
                else:
                    anchor_is_marker = True if (mode == "report_region" and anchor) else False
            r0 = r_anchor + (1 if anchor_is_marker else 0)

            # HARD-LOCK default: report_region templates usually have a header row at r0
            if "template_has_header" in t:
                template_has_header = bool(t.get("template_has_header"))
            else:
                # Only assume header row exists when anchor row is a marker
                template_has_header = True if (mode == "report_region" and anchor and anchor_is_marker) else False

            # Optional: clear anchor marker text (after resolve, before insert/write).
            if bool(t.get("anchor_clear", False)) and anchor:
                mode_clear = str(t.get("anchor_clear_mode") or "cell").lower()
                if mode_clear not in ("cell", "row"):
                    raise ValueError(f"excel_fill_from_file target={name}: invalid anchor_clear_mode={mode_clear}")
                if mode_clear == "cell":
                    _excel_set_cell_value(ws=ws, r=r_anchor, c=c0)
                else:
                    raw_w = t.get("anchor_clear_width", None)
                    try:
                        w_clear = int(raw_w) if raw_w is not None else 1
                    except Exception:
                        w_clear = 1
                    if w_clear > 0:
                        _excel_clear_row_segment(ws, r_anchor, c0, w_clear)

            # style copy options
            style_mode = str(t.get("style_mode") or "none").lower()
            if style_mode not in ("none", "copy_row"):
                raise ValueError(f"excel_fill_from_file target={name}: invalid style_mode={style_mode}")

            style_apply = str(t.get("style_apply") or "data").lower()
            if style_apply not in ("data", "header", "both"):
                raise ValueError(f"excel_fill_from_file target={name}: invalid style_apply={style_apply}")

            style_row_offset = int(t.get("style_row_offset", 2))
            clear_style_row = bool(t.get("clear_style_row", False))

            # row iterator
            def row_iter():
                if fmt in ("tsv", "csv"):
                    q = {
                        "minimal": csv.QUOTE_MINIMAL,
                        "all": csv.QUOTE_ALL,
                        "none": csv.QUOTE_NONE,
                        "nonnumeric": csv.QUOTE_NONNUMERIC,
                    }.get(str(quoting).strip().lower(), csv.QUOTE_MINIMAL)

                    with open(source_path, "r", encoding=encoding, newline="") as f:
                        reader = csv.reader(
                            f,
                            delimiter=delimiter,
                            quotechar=quotechar,
                            escapechar=escapechar,
                            doublequote=doublequote,
                            quoting=q,
                        )
                        for row in reader:
                            yield row

                elif fmt == "parquet":
                    try:
                        import pyarrow.parquet as pq
                    except Exception as e:
                        raise ValueError(
                            "parquet format requires optional dependency: pyarrow "
                            "(install aetherflow-core[parquet])"
                        ) from e
                    pf = pq.ParquetFile(source_path)
                    cols = pf.schema.names
                    if read_header and cols:
                        yield list(cols)
                    for batch in pf.iter_batches():
                        tbl = batch.to_pydict()
                        n = len(tbl[cols[0]]) if cols else 0
                        for i in range(n):
                            yield [tbl[c][i] for c in cols]
                else:
                    raise ValueError(f"Unsupported source_format: {fmt}")

            it = row_iter()

            # determine columns/header row
            columns = None
            header_row = None

            if fmt in ("tsv", "csv"):
                if read_header:
                    try:
                        header_row = next(it)
                        columns = list(header_row)
                    except StopIteration:
                        header_row = None
                        columns = list(provided_columns) if provided_columns else None
                else:
                    columns = list(provided_columns) if provided_columns else None

            elif fmt == "parquet":
                if read_header:
                    try:
                        header_row = next(it)
                        columns = list(header_row)
                    except StopIteration:
                        header_row = None
                        columns = None
                else:
                    columns = None

            if type_cast == "schema" and schema_map and not columns:
                type_cast = "none"

            will_write_header = (header_mode in ("override", "append")) and (header_row is not None)

            # HARD-LOCK: data starts below template header if template_has_header=True
            reserve_template_header_row = bool(template_has_header)
            data_start_row = r0 + (1 if (will_write_header or reserve_template_header_row) else 0)

            # report_region threshold pre-check (and used for insert count)
            row_count_data = None
            if mode == "report_region":
                fail_on_threshold = bool(t.get("fail_on_threshold", True))
                if provided_row_count is not None:
                    row_count_data = int(provided_row_count)
                else:
                    try:
                        row_count_data = int(
                            fast_count_rows(
                                source_path,
                                fmt,
                                include_header=read_header if fmt in ("csv", "tsv") else False,
                                count_mode=count_mode,
                                linefeed=linefeed,
                                encoding=encoding,
                                delimiter=delimiter,
                                quotechar=quotechar,
                                escapechar=escapechar,
                                doublequote=doublequote,
                                quoting=quoting,
                            )
                        )
                    except ParquetSupportMissing as e:
                        raise ValueError(str(e)) from e

                if fail_on_threshold and row_count_data is not None and row_count_data > rows_threshold:
                    raise ReportTooLargeError(
                        target_name=str(name),
                        source_path=str(source_path),
                        row_count=int(row_count_data),
                        rows_threshold=int(rows_threshold),
                    )

            # Insert rows (report_region only; preserves downstream anchors)
            if mode == "report_region" and insert == "below":
                n_data = int(row_count_data or 0)

                if reserve_template_header_row:
                    # HARD-LOCK: never insert at r0 when we must preserve template header row
                    if n_data > 0:
                        ws.insert_rows(data_start_row, amount=n_data)
                else:
                    # No reserved template header row: insert the full output block at r0
                    n_total = n_data + (1 if will_write_header else 0)
                    if n_total > 0:
                        ws.insert_rows(r0, amount=n_total)

            # Header clear (to remove old template header text) but keep formatting
            if will_write_header and header_mode in ("override", "append"):
                header_clear = bool(t.get("header_clear", True))
                header_clear_width = t.get("header_clear_width", None)

                w_new = len(header_row)
                if header_clear_width is None:
                    w_clear = w_new
                else:
                    try:
                        w_clear = int(header_clear_width)
                    except Exception:
                        w_clear = w_new

                if header_clear and w_clear > 0:
                    _excel_clear_row_segment(ws, r0, c0, w_clear)

            # Write header values (formatting stays unless header_style=style_row later)
            rows_written = 0
            width = 0

            if will_write_header:
                for j, v in enumerate(header_row):
                    _excel_set_cell_value(ws=ws, r=r0, c=c0+j, v=v)
                    width = max(width, j + 1)
                rows_written += 1

            # Write data rows
            data_row_index = 0
            for row in it:
                casted = _cast_row(row, columns=columns, type_cast=type_cast, schema_map=schema_map)
                width = max(width, len(casted))
                for j, v in enumerate(casted):
                    _excel_set_cell_value(ws=ws, r=data_start_row+data_row_index, c=c0+j, v=v)
                data_row_index += 1
            rows_written += data_row_index

            # Build style cache after width known
            style_cache = None
            style_row_idx = None
            if style_mode == "copy_row" and width > 0:
                style_row_idx = (r_anchor + style_row_offset) if anchor_is_marker else (r0 + style_row_offset)
                style_cache = _excel_build_style_row_cache(ws, style_row_idx, c0, width)

            # Apply styles:
            # - Header: only if header_style=style_row AND style_apply includes header
            # - Data: if style_apply includes data
            if style_cache is not None and rows_written > 0:
                if will_write_header and header_style == "style_row" and style_apply in ("header", "both"):
                    for j in range(min(width, len(style_cache))):
                        _excel_copy_cell_style(style_cache[j], ws.cell(row=r0, column=c0 + j))

                if data_row_index > 0 and style_apply in ("data", "both"):
                    for i in range(data_row_index):
                        rr = data_start_row + i
                        for j in range(min(width, len(style_cache))):
                            _excel_copy_cell_style(style_cache[j], ws.cell(row=rr, column=c0 + j))

                if clear_style_row and style_row_idx is not None:
                    _excel_clear_row_values(ws, style_row_idx, c0, width)

            # Overlap guard (post-write)
            # NOTE: Overlap only matters within the SAME sheet.
            # For data_sheet with insert=replace, we intentionally allow repeated writes to the same area.
            total_rows = (1 if (will_write_header or reserve_template_header_row) else 0) + data_row_index
            if not (mode == "data_sheet" and insert == "replace"):
                rect = (ws.title, r0, c0, r0 + max(0, total_rows - 1), c0 + max(0, width - 1))
                for prev in rects:
                    prev_sheet, pr1, pc1, pr2, pc2 = prev
                    if prev_sheet != ws.title:
                        continue  # only check overlap within same sheet
                    if _rect_intersects((pr1, pc1, pr2, pc2), rect[1:]):
                        raise ValueError(f"excel_fill_from_file targets overlap: name={name} sheet={ws.title} rect={rect} prev={prev}")

                rects.append(rect)

            # data_sheet niceties
            if mode == "data_sheet":
                if bool(t.get("freeze_panes", True)):
                    ws.freeze_panes = ws.cell(row=r0 + (1 if will_write_header else 0), column=c0)
                if bool(t.get("autofilter", True)) and rows_written > 0 and width > 0 and will_write_header:
                    from openpyxl.utils.cell import get_column_letter
                    ws.auto_filter.ref = (
                        f"{get_column_letter(c0)}{r0}:"
                        f"{get_column_letter(c0 + width - 1)}{r0 + rows_written - 1}"
                    )
                # HARD-LOCK: header_bold must NEVER replace template fonts.
                # It must only toggle bold=True on the existing font.
                if bool(t.get("header_bold", True)) and will_write_header:
                    from openpyxl.styles import Font
                    for j in range(width):
                        cell_obj = ws.cell(row=r0, column=c0 + j)
                        try:
                            base = cell_obj.font
                            f = _copy_obj(base) if base is not None else Font()
                            f.bold = True
                            cell_obj.font = f
                        except Exception as e:
                            log.warning(f"{name} header_bold failed {e}")
                            pass

            written.append(
                {
                    "name": name,
                    "sheet": ws.title,
                    "rows": int(total_rows),
                    "cols": int(width),
                    "mode": mode,
                    "insert": insert,
                    "header_mode": header_mode,
                    "header_style": header_style,
                    "style_mode": style_mode,
                    "start_cell": (r0, c0),
                    "template_has_header": bool(template_has_header),
                }
            )

            #_dbg_print_sheet_state(
            #    ws,
            #    label=f"after_target name={name} mode={mode} insert={insert} r_anchor={r_anchor} r0={r0} data_start={data_start_row}",
            #    focus_rows=[r_anchor, r0, data_start_row],
            #)

        #ws_report = wb["Report"] if "Report" in wb.sheetnames else None
        #if ws_report is not None:
        #    _dbg_print_sheet_state(ws_report, label="before_save", focus_rows=[])

        tmp = out_path.with_suffix(out_path.suffix + ".tmp")
        wb.save(tmp)
        os.replace(tmp, out_path)

        return {"output": str(out_path), "written": written}


def _dbg_find_anchors(ws, *, pattern_suffix="_ANCHOR", max_hits=50):
    hits = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.endswith(pattern_suffix):
                hits.append((v, cell.coordinate, cell.row, cell.column))
                if len(hits) >= max_hits:
                    return hits
    return hits

def _dbg_dump_range(ws, r1: int, r2: int, c1: int, c2: int):
    # returns a compact text table for quick console debugging
    lines = []
    for r in range(max(1, r1), r2 + 1):
        vals = []
        for c in range(max(1, c1), c2 + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                vals.append("")
            else:
                s = str(v)
                if len(s) > 30:
                    s = s[:27] + "..."
                vals.append(s)
        lines.append(f"R{r:>4}: " + " | ".join(vals))
    return "\n".join(lines)

def _dbg_print_sheet_state(ws, *, label: str, focus_rows: Iterable[int] | None = None):
    anchors = _dbg_find_anchors(ws)
    print(f"\n=== SHEET_STATE {label} sheet={ws.title} ===")
    print("anchors:", [(a[0], a[1]) for a in anchors])

    # dump around interesting rows (anchors + focus_rows)
    rows = set(focus_rows or [])
    for _, _, r, _ in anchors:
        rows.update([r - 2, r - 1, r, r + 1, r + 2])

    for r in sorted(x for x in rows if x >= 1)[:40]:
        # show first 8 columns (A..H)
        print(_dbg_dump_range(ws, r, r, 1, 8))


@register_step("with_lock")
class WithLock(Step):
    required_inputs = {"lock_key", "step"}

    def run(self) -> Dict[str, Any]:
        self.validate()
        key = self.inputs["lock_key"]
        ttl = int(self.inputs.get("ttl_seconds", 600))

        if not self.ctx.acquire_lock(key, ttl_seconds=ttl):
            raise RuntimeError(f"Lock not acquired: {key}")
        try:
            inner = self.inputs["step"]
            StepCls = get_step(inner["type"])
            step_id = inner.get("id", f"{self.id}_inner")
            inst = StepCls(step_id, inner.get("inputs") or {}, self.ctx, self.job_id)
            return inst.run()
        finally:
            self.ctx.release_lock(key)


def _safe_dest(base: Path, rel: str) -> Path:
    base = base.resolve()
    cand = (base / rel.lstrip("/\\")).resolve()
    try:
        cand.relative_to(base)
    except Exception:
        raise ValueError(f"dest path escapes dest_dir: {rel}")
    return cand


@register_step("smb_list_files")
class SMBListFiles(Step):
    required_inputs = {"resource", "remote_dir"}
    def run(self) -> Dict[str, Any]:
        """Best-effort recursive listing using the public SMB connector contract.

        IMPORTANT: bundle sync must not depend on private connector internals.

        SMB connectors are only required to expose:
          - list(remote_dir) -> list[RemoteFileMeta]

        We intentionally do not assume stat/scandir support here, so size/mtime
        may be unknown (None).
        """
        self.validate()
        smb = self.ctx.connectors[self.inputs["resource"]]
        pattern = self.inputs.get("pattern", "*")
        recursive = self.inputs.get("recursive", False)
        base_path = self.inputs["remote_dir"].rstrip("/\\") or "/"
        items: List[RemoteFileMeta] = []

        def _join(parent: str, name: str) -> str:
            parent = str(parent).rstrip("/\\")
            if not parent:
                return str(name)
            # Preserve "SHARE:/..." prefix semantics.
            if ":/" in parent:
                share, rest = parent.split(":/", 1)
                rest = rest.strip("/\\")
                joined = posixpath.join(rest, str(name)) if rest else str(name)
                return f"{share}:/{joined}"
            return posixpath.join(parent.replace("\\", "/"), str(name))

        def _walk(cur: str, rel_prefix: str, recursive: bool):
            try:
                entries = smb.list(cur)
                for e in entries or []:
                    if not e.name or e.name in {".", ".."}:
                        continue
                    # guard path
                    child = _join(cur, e.name)
                    rel = f"{rel_prefix}/{e.name}" if rel_prefix else str(e.name)
                    if e.is_dir and recursive:
                        if child and child != cur:   # avoid accidental self-loop
                            _walk(child, rel, recursive)
                        continue
                    if not fnmatch.fnmatch(e.name, pattern):
                        continue
                    items.append(replace(e, rel_path=rel))
            except Exception as e:
                raise ConnectorError(f"SMBListFiles list failed: {e}") from e

        _walk(base_path, "", recursive)
        count = len(items)
        min_count = int(self.inputs.get("min_count", 1))
        if count < min_count:
            return StepResult(
                status=STEP_SKIPPED,
                output={"has_data": False, "count": count, "files": []},
                reason=f"count({count}) < min_count({min_count})",
            )
        return StepResult(
            status=STEP_SUCCESS,
            output={"has_data": True, "count": count, "files": sorted(items, key=lambda x: x.rel_path)}
        )


@register_step("smb_download_files")
class SMBDownloadFiles(Step):
    required_inputs = {"resource", "files", "dest_dir"}
    def run(self) -> Dict[str, Any]:
        self.validate()
        smb = self.ctx.connectors[self.inputs["resource"]]
        job_dir = self.ctx.artifacts_dir(self.job_id)
        dest_dir = (job_dir / self.inputs["dest_dir"]).resolve()
        dest_dir.mkdir(parents=True, exist_ok=True)
        local_files = []
        for m in (self.inputs.get("files") or []):
            rel_path = m.get("rel_path") if isinstance(m, dict) else m.rel_path
            path = m.get("path") if isinstance(m, dict) else m.path
            dest = _safe_dest(dest_dir, rel_path)
            dest.parent.mkdir(parents=True, exist_ok=True)
            smb.download(path, str(dest))
            local_files.append(str(dest))
        return {"local_files": local_files, "dest_dir": str(dest_dir)}


@register_step("smb_delete_files")
class SMBDeleteFiles(Step):
    required_inputs = {"resource", "files"}
    def run(self) -> Dict[str, Any]:
        self.validate()
        smb = self.ctx.connectors[self.inputs["resource"]]
        for m in (self.inputs.get("files") or []):
            path = m.get("path") if isinstance(m, dict) else m.path
            smb.delete(path)
        return {"is_deleted": True}


@register_step("smb_upload_files")
class SMBUploadFiles(Step):
    required_inputs = {"resource", "local_files", "remote_dir"}
    def run(self) -> Dict[str, Any]:
        self.validate()
        smb = self.ctx.connectors[self.inputs["resource"]]
        remote_dir = self.inputs["remote_dir"]
        local_files: List[str] = self.inputs.get("local_files") or []
        uploaded = []
        for lf in local_files:
            lfp = Path(lf)
            if not lfp.is_absolute():
                lfp = (self.ctx.artifacts_dir(self.job_id) / lfp).resolve()
            smb.upload(lfp, f"{remote_dir}/{lfp.name}")
            uploaded.append(str(lfp))
        return {"uploaded": uploaded, "remote_dir": remote_dir}


@register_step("sftp_list_files")
class SFTPListFiles(Step):
    required_inputs = {"resource", "remote_dir"}
    def run(self) -> Dict[str, Any]:
        self.validate()
        sftp = self.ctx.connectors[self.inputs["resource"]]
        pattern = self.inputs.get("pattern", "*")
        recursive = self.inputs.get("recursive", False)
        base_path = self.inputs["remote_dir"].rstrip("/") or "/"
        items: List[RemoteFileMeta] = []

        def _walk(cur: str, recursive: bool):
            try:
                entries = sftp.list(cur)
                for e in entries:
                    # guard path
                    p = e.path or ""
                    if e.is_dir and recursive:
                        if p and p != cur:   # avoid accidental self-loop
                            _walk(p, recursive)
                        continue
                    if not fnmatch.fnmatch(e.name, pattern):
                        continue
                    rel = p[len(base_path) + 1 :] if (base_path != "/" and p.startswith(base_path + "/")) else (
                        p[1:] if (base_path == "/" and p.startswith("/")) else p
                    )
                    items.append(replace(e, rel_path=rel))
            except Exception as e:
                raise ConnectorError(f"SFTPListFiles list failed: {cur} {recursive} {e}") from e

        _walk(base_path, recursive)
        count = len(items)
        min_count = int(self.inputs.get("min_count", 1))
        if count < min_count:
            return StepResult(
                status=STEP_SKIPPED,
                output={"has_data": False, "count": count, "files": []},
                reason=f"count({count}) < min_count({min_count})",
            )
        return StepResult(
            status=STEP_SUCCESS,
            output={"has_data": True, "count": count, "files": sorted(items, key=lambda x: x.rel_path)}
        )


@register_step("sftp_download_files")
class SFTPDownloadFiles(Step):
    required_inputs = {"resource", "files", "dest_dir"}
    def run(self) -> Dict[str, Any]:
        self.validate()
        sftp = self.ctx.connectors[self.inputs["resource"]]
        job_dir = self.ctx.artifacts_dir(self.job_id)
        dest_dir = (job_dir / self.inputs["dest_dir"]).resolve()
        dest_dir.mkdir(parents=True, exist_ok=True)
        local_files = []
        for m in (self.inputs.get("files") or []):
            rel_path = m.get("rel_path") if isinstance(m, dict) else m.rel_path
            path = m.get("path") if isinstance(m, dict) else m.path
            dest = _safe_dest(dest_dir, rel_path)
            dest.parent.mkdir(parents=True, exist_ok=True)
            sftp.download(path, str(dest))
            local_files.append(str(dest))
        return {"local_files": local_files, "dest_dir": str(dest_dir)}


@register_step("sftp_delete_files")
class SFTPDeleteFiles(Step):
    required_inputs = {"resource", "files"}
    def run(self) -> Dict[str, Any]:
        self.validate()
        sftp = self.ctx.connectors[self.inputs["resource"]]
        for m in (self.inputs.get("files") or []):
            path = m.get("path") if isinstance(m, dict) else m.path
            sftp.delete(path)
        return {"is_deleted": True}


@register_step("sftp_upload_files")
class SFTPUploadFiles(Step):
    required_inputs = {"resource", "items", "remote_dir"}

    def run(self) -> Dict[str, Any]:
        self.validate()
        sftp = self.ctx.connectors[self.inputs["resource"]]
        items = self.inputs["items"]
        remote_dir = self.inputs["remote_dir"].rstrip("/")

        p = self.inputs.get("parallelism") or {}
        enabled = bool(p.get("enabled", True))
        workers = int(p.get("workers", 8))
        fail_fast = bool(p.get("fail_fast", True))

        manifest = self.ctx.manifests_dir(self.job_id) / f"{self.id}.manifest.json"
        done = set()

        if manifest.exists():
            try:
                done = {x["id"] for x in json.loads(manifest.read_text(encoding="utf-8"))}
            except Exception:
                done = set()

        def one(it):
            local = Path(it) if isinstance(it, str) else Path(it["file"])
            item_id = local.name if isinstance(it, str) else (it.get("id") or local.name)
            remote_name = local.name if isinstance(it, str) else (it.get("remote_name") or local.name)
            if item_id in done:
                return {"id": item_id, "skipped": True}
            remote_path = f"{remote_dir}/{remote_name}"
            sftp.upload(str(local), remote_path)
            return {"id": item_id, "file": str(local), "remote_path": remote_path}

        outs = run_thread_pool(items, one, workers=workers, fail_fast=fail_fast) if enabled and len(items) > 1 else [one(x) for x in items]

        tmp = manifest.with_suffix(".tmp")
        tmp.write_text(json.dumps(outs, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(tmp, manifest)
        return {"manifest": str(manifest), "count": len(outs)}


@register_step("check_items")
class CheckItems(Step):
    """Gate step: check that a list of items is non-empty.

    This is a small, generic probe step for demos and simple flows.

    Inputs:
      - items: list | string (comma-separated)
      - min_count: int (default 1)

    Output:
      - has_data: bool
      - count: int
      - reason: string (when skipped)
    """

    required_inputs = {"items"}

    def run(self) -> Dict[str, Any]:
        self.validate()
        items = self.inputs.get("items")
        if isinstance(items, str):
            items = [x.strip() for x in items.split(",") if x.strip()]
        if not isinstance(items, list):
            raise ValueError("check_items expects 'items' to be a list or a comma-separated string")

        min_count = int(self.inputs.get("min_count", 1))
        count = len(items)
        if count < min_count:
            return StepResult(
                status=STEP_SKIPPED,
                output={"has_data": False, "count": count},
                reason=f"count({count}) < min_count({min_count})",
            )
        return StepResult(status=STEP_SUCCESS, output={"has_data": True, "count": count})


@register_step("mail_send")
class MailSend(Step):
    """Send an email via a configured mail resource.

    Inputs:
      - resource: mail resource name (kind=mail, driver=smtp)
      - to: string or list of strings
      - subject: string
      - body: plaintext body OR html body (when html=true)
      - html: bool (default false)
      - text: optional plaintext fallback when html=true
      - cc: optional string/list
      - bcc: optional string/list
      - from_addr: optional override
    """

    required_inputs = {"resource", "to", "subject", "body"}

    def run(self) -> Dict[str, Any]:
        self.validate()
        mail = self.ctx.connectors.mail(self.inputs["resource"]) if hasattr(self.ctx.connectors, "mail") else self.ctx.connectors[self.inputs["resource"]]

        subject = str(self.inputs["subject"])
        body = str(self.inputs["body"])
        html = bool(self.inputs.get("html", False))

        from_addr = self.inputs.get("from_addr")
        to = _as_list(self.inputs["to"])
        cc = _as_list(self.inputs.get("cc"))
        bcc = _as_list(self.inputs.get("bcc"))

        if html:
            text = self.inputs.get("text")
            mail.send_html(to=to, subject=subject, html=body, text=text, cc=cc, bcc=bcc, from_addr=from_addr)
        else:
            mail.send_plaintext(to=to, subject=subject, body=body, cc=cc, bcc=bcc, from_addr=from_addr)

        return {"sent": True, "to": to, "subject": subject, "html": html}


@register_step("external.process")
class ExternalProcess(Step):
    """Run an external OS-level process.

    This step is the ops-grade bridge to big-data / ML engines.

    Features:
      - stdout/stderr handling: inherit | capture | file | discard
      - timeout + graceful shutdown (SIGTERM -> kill)
      - retry policy: exit codes, timeout
      - success validation: required files/globs, forbidden files, marker
      - idempotency: marker, atomic output directory

    Notes:
      - Relative paths are resolved against the job artifacts directory.
      - `{{run_id}}` in paths is rendered with the current run id.
    """

    required_inputs = {"command"}

    def run(self) -> Dict[str, Any]:
        self.validate()

        settings = getattr(self.ctx, "settings", None)
        logger = getattr(self.ctx, "log", None)

        command = self.inputs.get("command")
        args = self.inputs.get("args")
        shell = bool(self.inputs.get("shell", False))

        # Prefer list-form for predictability
        if isinstance(command, list):
            cmd: list[str] = [str(x) for x in command]
        elif isinstance(command, str):
            cmd = [command]
        else:
            raise ValueError("external.process inputs.command must be a string or list")

        if args is not None:
            if not isinstance(args, list):
                raise ValueError("external.process inputs.args must be a list")
            cmd.extend([str(x) for x in args])

        cwd = self.inputs.get("cwd")
        cwd_path = str(_resolve_path(self.ctx, self.job_id, cwd)) if cwd else None

        base_env = getattr(self.ctx, "env", None) or {}
        env = dict(base_env) if bool(self.inputs.get("inherit_env", True)) else {}
        env.setdefault("AETHERFLOW_FLOW_ID", getattr(self.ctx, "flow_id", ""))
        env.setdefault("AETHERFLOW_RUN_ID", getattr(self.ctx, "run_id", ""))
        for k, v in (self.inputs.get("env") or {}).items():
            env[str(k)] = "" if v is None else str(v)

        timeout = self.inputs.get("timeout_seconds")
        timeout = float(timeout) if timeout is not None else None
        kill_grace = int(self.inputs.get("kill_grace_seconds", 15))

        # --- logging config ---
        log_cfg = self.inputs.get("log") or {}
        stdout_mode = (log_cfg.get("stdout") or "inherit").lower()
        stderr_mode = (log_cfg.get("stderr") or "inherit").lower()

        max_cap = int(log_cfg.get("max_capture_kb", 1024)) * 1024
        file_path = log_cfg.get("file_path")
        log_file_path = _resolve_path(self.ctx, self.job_id, file_path) if file_path else None

        out_cap = _Capture(mode="capture", max_bytes=max_cap)
        err_cap = _Capture(mode="capture", max_bytes=max_cap)

        file_handle = None
        def _io_for(mode: str):
            nonlocal file_handle
            if mode == "discard":
                return subprocess.DEVNULL
            if mode == "inherit":
                return None
            if mode == "file":
                if not log_file_path:
                    raise ValueError("log.file_path is required when stdout/stderr mode is 'file'")
                log_file_path.parent.mkdir(parents=True, exist_ok=True)
                if file_handle is None:
                    file_handle = open(log_file_path, "ab")
                return file_handle
            if mode == "capture":
                return subprocess.PIPE
            raise ValueError(f"Unsupported log mode: {mode}")

        # --- idempotency / success criteria ---
        idem = self.inputs.get("idempotency") or {}
        strategy = (idem.get("strategy") or "none").lower()

        success_spec = self.inputs.get("success") or {}
        exit_codes = [int(x) for x in _as_list(success_spec.get("exit_codes"))] or [0]

        if strategy == "marker":
            marker_path = idem.get("marker_path") or success_spec.get("marker_file")
            if marker_path:
                mp = _resolve_path(self.ctx, self.job_id, marker_path)
                if mp.exists():
                    ok, _reason = _check_success(ctx=self.ctx, job_id=self.job_id, spec=success_spec)
                    if ok:
                        return StepResult(status=STEP_SKIPPED, output={"skipped": True, "marker": str(mp)}, reason="marker_present")

        tmp_out = None
        final_out = None
        if strategy == "atomic_dir":
            tmp_out = idem.get("temp_output_dir")
            final_out = idem.get("final_output_dir")
            if not tmp_out or not final_out:
                raise ValueError("atomic_dir requires temp_output_dir and final_output_dir")
            tmp_p = _resolve_path(self.ctx, self.job_id, tmp_out)
            if tmp_p.exists():
                shutil.rmtree(tmp_p, ignore_errors=True)
            tmp_p.mkdir(parents=True, exist_ok=True)
            env.setdefault("AETHERFLOW_OUTPUT_DIR", str(tmp_p))

        retry = self.inputs.get("retry") or {}
        max_attempts = int(retry.get("max_attempts", 1))
        backoff = float(retry.get("backoff_seconds", 0) or 0)
        mult = float(retry.get("backoff_multiplier", 1.0) or 1.0)
        max_backoff = float(retry.get("max_backoff_seconds", 0) or 0)
        retry_on_exit = [int(x) for x in _as_list(retry.get("retry_on_exit_codes"))]
        retry_on_timeout = bool(retry.get("retry_on_timeout", False))

        fields = {"job_id": self.job_id, "step_id": self.id, "step_type": "external.process"}

        def _sleep(sec: float) -> None:
            if sec <= 0:
                return
            time.sleep(sec)

        last_err: Optional[str] = None

        for attempt in range(1, max_attempts + 1):
            # backoff before retries
            if attempt > 1 and backoff > 0:
                sleep_s = backoff * (mult ** (attempt - 2))
                if max_backoff and sleep_s > max_backoff:
                    sleep_s = max_backoff
                _sleep(sleep_s)

            stdout_io = _io_for(stdout_mode)
            stderr_io = _io_for(stderr_mode)

            # IMPORTANT: reset capture per attempt
            out_cap = _Capture(mode="capture", max_bytes=max_cap)
            err_cap = _Capture(mode="capture", max_bytes=max_cap)

            threads: list[threading.Thread] = []

            if logger and settings:
                log_event(logger, settings=settings, level=20, event="external_process_start", attempt=attempt, command=cmd, **fields)

            p = None
            try:
                if shell:
                    import shlex
                    cmd_str = " ".join(shlex.quote(x) for x in cmd)
                    popen_args = cmd_str
                    log.warning("Shell=True - POSIX quoting used; Windows behavior may differ")
                else:
                    popen_args = cmd
                p = subprocess.Popen(popen_args, shell=shell, cwd=cwd_path, env=env, stdout=stdout_io, stderr=stderr_io)
                if stdout_mode == "capture" and p.stdout is not None:
                    th = threading.Thread(
                        target=_stream_reader,
                        args=(p.stdout, out_cap, logger),
                        kwargs={
                            "settings": settings,
                            "event_prefix": "external_process_stdout",
                            "fields": fields,
                            "level": 20,
                        },
                        daemon=True,
                    )
                    th.start()
                    threads.append(th)
                if stderr_mode == "capture" and p.stderr is not None:
                    th = threading.Thread(
                        target=_stream_reader,
                        args=(p.stderr, err_cap, logger),
                        kwargs={
                            "settings": settings,
                            "event_prefix": "external_process_stderr",
                            "fields": fields,
                            "level": 30,
                        },
                        daemon=True,
                    )
                    th.start()
                    threads.append(th)

                timed_out = False
                try:
                    p.wait(timeout=timeout)
                except subprocess.TimeoutExpired:
                    timed_out = True
                    try:
                        p.send_signal(signal.SIGTERM)
                    except Exception:
                        pass
                    try:
                        p.wait(timeout=kill_grace)
                    except Exception:
                        try:
                            p.kill()
                        except Exception:
                            pass
                        try:
                            p.wait(timeout=5)
                        except Exception:
                            pass

                rc = p.returncode

                if timed_out:
                    last_err = "timeout"
                    if logger and settings:
                        log_event(logger, settings=settings, level=40, event="external_process_timeout", attempt=attempt, **fields)
                    if retry_on_timeout and attempt < max_attempts:
                        continue
                    raise TimeoutError(f"external.process timed out after {timeout}s")

                if rc not in exit_codes:
                    last_err = f"exit_code:{rc}"
                    if logger and settings:
                        log_event(logger, settings=settings, level=40, event="external_process_exit", attempt=attempt, exit_code=rc, **fields)
                    if rc in retry_on_exit and attempt < max_attempts:
                        continue
                    raise RuntimeError(f"external.process failed with exit code {rc}\n{cmd} - {popen_args}\n{err_cap.text()}")

                # finalize atomic output directory
                if strategy == "atomic_dir":
                    tmp_p = _resolve_path(self.ctx, self.job_id, tmp_out)
                    final_p = _resolve_path(self.ctx, self.job_id, final_out)
                    if bool(idem.get("atomic_rename", True)):
                        if final_p.exists():
                            shutil.rmtree(final_p, ignore_errors=True)
                        os.replace(tmp_p, final_p)
                    else:
                        final_p.mkdir(parents=True, exist_ok=True)
                        for item in tmp_p.iterdir():
                            shutil.move(str(item), str(final_p / item.name))
                        shutil.rmtree(tmp_p, ignore_errors=True)

                ok, reason = _check_success(ctx=self.ctx, job_id=self.job_id, spec=success_spec)
                if not ok:
                    last_err = f"invalid_output:{reason}"
                    raise RuntimeError(f"external.process outputs invalid: {reason}")

                out: Dict[str, Any] = {"exit_code": rc, "attempts": attempt}
                if stdout_mode == "capture":
                    out["stdout"] = out_cap.text()
                if stderr_mode == "capture":
                    out["stderr"] = err_cap.text()
                if log_file_path and (stdout_mode == "file" or stderr_mode == "file"):
                    out["log_file"] = str(log_file_path)

                for k, v in (self.inputs.get("outputs") or {}).items():
                    out[str(k)] = v

                if logger and settings:
                    log_event(logger, settings=settings, level=20, event="external_process_success", attempt=attempt, exit_code=rc, **fields)

                return out
            finally:
                # close pipes first to unblock _stream_reader.readline()
                if p is not None:
                    try:
                        if getattr(p, "stdout", None) is not None:
                            p.stdout.close()
                    except Exception:
                        pass
                    try:
                        if getattr(p, "stderr", None) is not None:
                            p.stderr.close()
                    except Exception:
                        pass
                for th in threads:
                    try:
                        th.join()
                    except Exception:
                        pass
                # close file handle once
                try:
                    if file_handle is not None:
                        file_handle.close()
                except Exception:
                    pass

        raise RuntimeError(last_err or "external_process_failed")


def _iter_input_paths(ctx, job_id: str, items: list[Any], *, base_dir: Path, recursive: bool) -> list[Path]:
    """Collect input files for zipping.

    - strings can be file/dir paths or glob patterns
    - dict items can specify {"path": ..., "glob": ...}
    """
    out: list[Path] = []

    def add_path(p: Path) -> None:
        p = p.resolve()
        if p.is_dir():
            if recursive:
                for fp in p.rglob("*"):
                    if fp.is_file():
                        out.append(fp)
        elif p.is_file():
            out.append(p)

    for it in items:
        if isinstance(it, dict):
            if "path" in it:
                add_path(_resolve_path(ctx, job_id, str(it["path"])))
            elif "glob" in it:
                g = str(it["glob"])
                gp = _resolve_path(ctx, job_id, g)
                for fp in gp.parent.glob(gp.name):
                    add_path(fp)
            else:
                raise ValueError("zip inputs item dict must have 'path' or 'glob'")
        else:
            s = str(it)
            # treat as glob if contains wildcard
            if any(ch in s for ch in ["*", "?", "["]):
                gp = _resolve_path(ctx, job_id, s)
                for fp in gp.parent.glob(gp.name):
                    add_path(fp)
            else:
                add_path(_resolve_path(ctx, job_id, s))

    # de-dup, keep deterministic order
    uniq = sorted({p for p in out})

    # Ensure all files are under base_dir for stable archive names.
    safe: list[Path] = []
    for fp in uniq:
        try:
            fp.relative_to(base_dir)
        except Exception:
            raise ValueError(f"zip input '{fp}' is outside base_dir '{base_dir}'")
        safe.append(fp)
    return safe


def _has_cmd(cmd: str) -> bool:
    return shutil.which(cmd) is not None


def _get_archive_connector(ctx: Any, resource_name: str):
    """Resolve an archive connector from ctx.connectors.

    Works with both Connectors manager (preferred) and legacy dict-style.
    """
    conns = ctx.connectors
    if hasattr(conns, "archive"):
        return conns.archive(resource_name)
    if hasattr(conns, "get"):
        return conns.get(kind="archive", name=resource_name)
    return conns[resource_name]


@register_step("zip")
class ZipCreate(Step):
    """Create a .zip archive.

    Inputs:
      - dest_path: path to .zip (relative -> job artifacts dir)
      - items: list of paths/globs (relative -> job artifacts dir). May include dirs.
      - src_dir: optional base dir for archive names (default: job artifacts dir)
      - recursive: bool (default true) when items include dirs
      - resource / connector: optional archive resource name (kind: archive).
        If omitted, Aetherflow will try to use a resource named `archive_default`.
        If that also does not exist, it falls back to stdlib zipfile for no-password,
        and to pyzipper/OS tools for password when available.
      - password: optional password. Encryption depends on archive driver (pyzipper=AES, os/pyminizip=ZipCrypto legacy).
      - overwrite: bool (default true)

    Notes:
      - Recommended: define `resources.archive_default` with `kind: archive` and `driver: pyzipper`.
      - For password-protected archives, use `archive:pyzipper` (AES). For ZipCrypto compatibility, use `archive:pyminizip` or `archive:os`.
      - `archive:os` uses system tools (`zip`/`unzip`) when you want to depend on the OS.
      - `archive:external` can call out to tools like 7z.
    """

    required_inputs = {"dest_path", "items"}

    def run(self) -> Dict[str, Any]:
        self.validate()

        out_path = _resolve_path(self.ctx, self.job_id, str(self.inputs["dest_path"]))

        items = self.inputs.get("items")
        if not isinstance(items, list):
            raise ValueError("zip inputs.items must be a list")

        base_dir_in = self.inputs.get("src_dir")
        base_dir = _resolve_path(self.ctx, self.job_id, str(base_dir_in)) if base_dir_in else self.ctx.artifacts_dir(self.job_id)
        base_dir = base_dir.resolve()

        recursive = bool(self.inputs.get("recursive", True))
        overwrite = bool(self.inputs.get("overwrite", True))
        password = self.inputs.get("password")

        files = _iter_input_paths(self.ctx, self.job_id, items, base_dir=base_dir, recursive=recursive)
        if not files:
            return StepResult(status=STEP_SKIPPED, output={"output": str(out_path), "count": 0}, reason="no_files")

        # Resolve archive connector
        resource = self.inputs.get("resource") or self.inputs.get("connector")
        # Convention: archive_default
        if not resource and hasattr(self.ctx.connectors, "resources"):
            if "archive_default" in getattr(self.ctx.connectors, "resources", {}):
                resource = "archive_default"

        arch = None
        if resource:
            arch = _get_archive_connector(self.ctx, str(resource))
        else:
            if _is_enterprise(self.ctx):
                raise RuntimeError("Zip/Unzip requires declared archive resource in enterprise mode")
            # Back-compat fallback (no resource provided)
            from aetherflow.core.registry.connectors import REGISTRY
            if password:
                try:
                    import pyzipper  # noqa: F401
                    arch = REGISTRY.create(name="_adhoc", kind="archive", driver="pyzipper", config={}, options={}, ctx=self.ctx)
                except Exception:
                    if _has_cmd("zip") and _has_cmd("unzip"):
                        arch = REGISTRY.create(name="_adhoc", kind="archive", driver="os", config={}, options={}, ctx=self.ctx)
                    else:
                        raise RuntimeError(
                            "Encrypted zip requires an archive connector. Provide inputs.resource (kind=archive), "
                            "or define a resource named 'archive_default'. Recommended: kind=archive driver=pyzipper."
                        )
            else:
                arch = REGISTRY.create(name="_adhoc", kind="archive", driver="zipfile", config={}, options={}, ctx=self.ctx)

        compression = (self.inputs.get("compression") or "deflated")
        out = arch.create_zip(
            output=out_path,
            files=files,
            base_dir=base_dir,
            password=str(password) if password else None,
            compression=str(compression),
            overwrite=overwrite,
        )
        out.update({"src_dir": str(base_dir)})
        return out


@register_step("unzip")
class ZipExtract(Step):
    """Extract multiple .zip archives.

    Inputs:
      - src_dir: compute relative path
      - archives: path to zip file (relative -> job artifacts dir)
      - dest_dir: destination directory (relative -> job artifacts dir)
      - resource / connector: optional archive resource name (kind: archive).
        If omitted, Aetherflow will try to use a resource named `archive_default`.
        If that also does not exist, it falls back to a built-in adhoc connector
        (zipfile for non-encrypted, pyzipper/os for encrypted if available).
      - password: optional password. Driver behavior: pyzipper=AES, zipfile can read ZipCrypto, os depends on OS tools.
      - overwrite: bool (default true)
      - members: optional list of specific member paths to extract

    Notes:
      - Prefer `archive.pyzipper` for cross-platform encrypted zips.
      - `archive.zipfile` cannot write encrypted zips (but can read ZipCrypto).
    """

    required_inputs = {"archives", "dest_dir"}

    def run(self) -> Dict[str, Any]:
        self.validate()

        extracted = []
        archives = self.inputs["archives"]
        if archives is None:
            return {}
        if isinstance(archives, (str, Path)):
            archives = [str(archives)]
        if not isinstance(archives, list):
            raise ValueError(f"Archives must be a list of Path {archives} ")
        if not archives:
            return {}

        src_dir_in = self.inputs.get("src_dir")
        src_dir = _resolve_path(self.ctx, self.job_id, src_dir_in) if src_dir_in else self.ctx.artifacts_dir(self.job_id)
        dest_dir = _resolve_path(self.ctx, self.job_id, str(self.inputs["dest_dir"]))
        dest_dir.mkdir(parents=True, exist_ok=True)

        overwrite = bool(self.inputs.get("overwrite", True))
        private_zip_folder = bool(self.inputs.get("private_zip_folder", False))
        password = self.inputs.get("password")
        members = self.inputs.get("members")
        if members is not None and not isinstance(members, list):
            raise ValueError("unzip inputs.members must be a list when provided")

        # Resolve archive connector
        resource = self.inputs.get("resource") or self.inputs.get("connector")
        if not resource and hasattr(self.ctx.connectors, "resources"):
            if "archive_default" in getattr(self.ctx.connectors, "resources", {}):
                resource = "archive_default"

        arch = None
        if resource:
            arch = _get_archive_connector(self.ctx, str(resource))
        else:
            if _is_enterprise(self.ctx):
                raise RuntimeError("Zip/Unzip requires declared archive resource in enterprise mode")
            # Back-compat fallback (no resource provided)
            from aetherflow.core.registry.connectors import REGISTRY
            if password:
                try:
                    import pyzipper  # noqa: F401
                    arch = REGISTRY.create(name="_adhoc", kind="archive", driver="pyzipper", config={}, options={}, ctx=self.ctx)
                except Exception:
                    if _has_cmd("zip") and _has_cmd("unzip"):
                        arch = REGISTRY.create(name="_adhoc", kind="archive", driver="os", config={}, options={}, ctx=self.ctx)
                    else:
                        raise RuntimeError(
                            "Encrypted unzip requires an archive connector. Provide inputs.resource (kind=archive), "
                            "or define a resource named 'archive_default'. Recommended: kind=archive driver=pyzipper."
                        )
            else:
                arch = REGISTRY.create(name="_adhoc", kind="archive", driver="zipfile", config={}, options={}, ctx=self.ctx)

        for a in archives:
            ap = _resolve_path(self.ctx, self.job_id, a)
            if not ap.exists() or ap.is_dir():
                continue

            try:
                rel = ap.relative_to(src_dir)         # e.g. cde/fgh.zip
                if private_zip_folder:
                    rel_no_ext = rel.with_suffix("")
                    out_dir = dest_dir / rel_no_ext
                else:
                    out_dir = dest_dir / rel.parent         # e.g. dest/cde
            except ValueError:
                out_dir = dest_dir / Path(ap.name)          # fallback if not under scan_root

            out_dir.mkdir(parents=True, exist_ok=True)
            eap = arch.extract_zip(
                archive=ap,
                dest_dir=out_dir,
                password=str(password) if password else None,
                overwrite=overwrite,
                members=[str(m) for m in members] if members else None,
            )
            extracted.append(eap)

        return {"unzipped": extracted, "dest_dir": str(dest_dir)}
